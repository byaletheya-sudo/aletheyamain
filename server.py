from flask import Flask, request, jsonify, send_from_directory, send_file, session, redirect
from openai import OpenAI
import base64
import os
import re
import io
import csv
import json
import hmac
import hashlib
import time
import zipfile
import secrets
import threading
import rx_form
import urllib.request
import urllib.parse
import urllib.error
import socket
import ipaddress
from datetime import timedelta
from html import unescape as _html_unescape

app = Flask(__name__)

# Session signing key. NEVER hardcode a real one — this repo is public, and a known
# key lets anyone forge a "logged-in" cookie and skip the password entirely. Use the
# SECRET_KEY env var in production; otherwise fall back to a random per-process key
# (sessions simply won't survive a restart, which just means logging in again).
app.secret_key = os.environ.get("SECRET_KEY") or secrets.token_hex(32)

# Harden the session cookie.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,    # not readable by JS (an XSS can't steal the session)
    SESSION_COOKIE_SECURE=True,      # only sent over HTTPS (Railway/custom domains are HTTPS)
    SESSION_COOKIE_SAMESITE="Lax",   # basic CSRF mitigation
)

# Password gate for the whole app. Env-only in practice: the fallback is random, so
# the value that used to be committed here no longer opens the door. Set APP_PASSWORD
# in Railway → Variables.
APP_PASSWORD = os.environ.get("APP_PASSWORD") or secrets.token_urlsafe(24)
if not os.environ.get("APP_PASSWORD"):
    print("[security] APP_PASSWORD is not set — using a random password. Set it in Railway to log in.")

# Second-tier gate for the restricted tools (Lease Ad — TEST + Deal Hub). Override
# with RESTRICTED_PASSWORD in Railway (recommended — this repo is public).
RESTRICTED_PASSWORD = os.environ.get("RESTRICTED_PASSWORD") or "notforyou"
# Lease Ad Generator is no longer second-password-protected, so its shared photo/parse
# endpoints (/carsxe-*, /carvector-*, /bulk-parse) are open past the primary login like
# the Sold/Review tools. The Deal Hub / Invoice / Desking DATA endpoints stay gated.
RESTRICTED_API = ("/deal-parse", "/deal-search", "/deals", "/contacts", "/published", "/verify-", "/invoices", "/desk-parse", "/desk-programs", "/copilot-kb")

# Aletheya Toolbox — a separate, non-Nova workspace with its OWN password gate.
# Override with TOOLBOX_PASSWORD in the host env (recommended — this repo is public).
TOOLBOX_PASSWORD = os.environ.get("TOOLBOX_PASSWORD") or "arvin"

# Nova Admins — the back-office suite (deals / payroll / tasks / notes). Its OWN gate
# (session["admin_ok"]), separate from the Nova workspace login. FAIL-CLOSED: there is
# deliberately NO default password — this repo is public and the suite holds financial,
# commission and client data. Until NOVA_ADMIN_PASSWORD is set in the host env, the
# suite cannot be unlocked at all.
NOVA_ADMIN_PASSWORD = os.environ.get("NOVA_ADMIN_PASSWORD", "").strip()
if not NOVA_ADMIN_PASSWORD:
    print("[security] NOVA_ADMIN_PASSWORD is NOT set — Nova Admins is LOCKED until you set it.")

# Optional shared-secret token for headless writes to Nova Admins (seeding the live
# volume, the nightly Garage sync). When set, an X-Nova-Token header that matches is
# accepted in place of an admin browser session. No default; short tokens are refused.
NOVA_ADMIN_TOKEN = os.environ.get("NOVA_ADMIN_TOKEN", "").strip()
if NOVA_ADMIN_TOKEN and len(NOVA_ADMIN_TOKEN) < 20:
    print("[security] NOVA_ADMIN_TOKEN is shorter than 20 chars — ignoring it. Use a long random secret.")
    NOVA_ADMIN_TOKEN = ""

# Admin session policy (Edgar's call): stay signed in for as long as the server is up —
# NO idle or absolute time-out. The only automatic sign-out is a SERVER REFRESH (restart):
# each login is stamped with the current process's boot id, and when the server restarts
# that id changes so stale sessions are dropped. Explicit Lock/logout still signs out.
SERVER_BOOT_ID = secrets.token_hex(8)
app.permanent_session_lifetime = timedelta(days=30)   # persisted-cookie lifetime (survives browser restarts)

# Stricter throttle for the admin gate than the general site login.
_ADMIN_MAX = 5
_ADMIN_WINDOW = 600   # seconds

# ---- Nova Admins user accounts (real per-user logins) ----
# The owners. Names/roles are safe to keep here; password HASHES live in the
# gitignored store under "users", so credentials never touch this public repo.
NOVA_USERS = [
    {"id": "edgar", "name": "Edgar", "role": "owner"},
    {"id": "nema",  "name": "Nema",  "role": "owner"},
    {"id": "arvin", "name": "Arvin", "role": "owner"},
]
NOVA_USER_IDS = {u["id"] for u in NOVA_USERS}
NOVA_ACCOUNT_ADMIN = "edgar"   # only Edgar can set OTHER people's passwords


def _hash_pw(pw):
    salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", pw.encode("utf-8"), bytes.fromhex(salt), 200_000).hex()
    return "pbkdf2$200000$" + salt + "$" + dk


def _check_pw(pw, stored):
    try:
        _a, iters, salt, dk = (stored or "").split("$")
        calc = hashlib.pbkdf2_hmac("sha256", pw.encode("utf-8"), bytes.fromhex(salt), int(iters)).hex()
        return hmac.compare_digest(calc, dk)
    except Exception:
        return False


def _nova_roster():
    """{id: record} merged from the store, always including the 3 owners."""
    store = _nova_load()
    by_id = {u.get("id"): dict(u) for u in store.get("users", [])
             if isinstance(u, dict) and u.get("id") in NOVA_USER_IDS}
    for r in NOVA_USERS:
        rec = by_id.setdefault(r["id"], {"id": r["id"], "pass": None})
        rec["name"], rec["role"] = r["name"], r["role"]
    return by_id


def _nova_set_password(uid, pw):
    if uid not in NOVA_USER_IDS or not pw:
        return False
    with _NOVA_LOCK:
        store = _nova_load()
        users = store.setdefault("users", [])
        rec = next((u for u in users if u.get("id") == uid), None)
        if not rec:
            base = next(r for r in NOVA_USERS if r["id"] == uid)
            rec = {"id": uid, "name": base["name"], "role": base["role"]}
            users.append(rec)
        rec["pass"] = _hash_pw(pw)
        _nova_write(store)
    return True


def _nova_current_user():
    uid = session.get("nova_user")
    # Legacy sessions (created before per-user logins) carry admin_ok but no identity.
    # They were authenticated with the master admin password — Edgar's bootstrap
    # credential — so treat them as Edgar instead of locking him out of Team/agent.
    if not uid and session.get("admin_ok"):
        uid = NOVA_ACCOUNT_ADMIN
    r = next((x for x in NOVA_USERS if x["id"] == uid), None)
    return {"id": r["id"], "name": r["name"], "role": r["role"]} if r else None


def _audit(ev, **kw):
    """Append-only audit trail for the Nova Admins suite (on the data volume,
    never in git). One JSON line per event: who, what, when, from where."""
    try:
        rec = {"ts": time.strftime("%Y-%m-%d %H:%M:%S"), "ip": _client_ip(), "ev": ev}
        rec.update(kw)
        with open(os.path.join(GENERATED_DIR, "nova_admins_audit.jsonl"), "a", encoding="utf-8") as f:
            f.write(json.dumps(rec) + "\n")
    except Exception:
        pass

# Tiny in-memory brute-force throttle: max attempts per IP per window.
_LOGIN_FAILS = {}
_LOGIN_MAX = 8
_LOGIN_WINDOW = 300   # seconds


def _client_ip():
    xff = request.headers.get("X-Forwarded-For", "")
    return (xff.split(",")[0].strip() if xff else (request.remote_addr or "?"))


def login_page(message="", next_path="/novauto"):
    msg = f'<p class="err">{message}</p>' if message else ""
    safe_next = next_path if next_path.startswith("/") else "/novauto"
    nxt = f'<input type="hidden" name="next" value="{safe_next}">'
    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0"><title>nova.byAletheya</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:#05080d; color:#fff;
    min-height:100vh; display:flex; align-items:center; justify-content:center;
    background-image: radial-gradient(70% 50% at 50% 35%, rgba(52,118,222,.30), rgba(8,16,30,0) 70%); }}
  .card {{ width:340px; max-width:90vw; background:#101319; border:1px solid #232733; border-radius:16px; padding:34px 28px; text-align:center; box-shadow:0 30px 80px rgba(0,0,0,.5); }}
  .brand {{ font-size:1.5rem; font-weight:800; letter-spacing:-.5px; margin-bottom:4px; }}
  .brand span {{ color:#4a9eff; }}
  .sub {{ color:#6b7280; font-size:.72rem; letter-spacing:.12em; text-transform:uppercase; margin-bottom:24px; }}
  input {{ width:100%; background:#0a0d12; border:1px solid #2a2f3a; border-radius:9px; padding:13px 15px; color:#fff; font-size:.95rem; outline:none; }}
  input:focus {{ border-color:#4a9eff; }}
  button {{ width:100%; margin-top:12px; padding:13px; border:none; border-radius:9px; background:#3a8eef; color:#fff; font-weight:600; font-size:.95rem; cursor:pointer; }}
  button:hover {{ background:#327fe0; }}
  .err {{ color:#ff6b6b; font-size:.82rem; margin-bottom:12px; }}
  .legal {{ margin-top:20px; padding-top:16px; border-top:1px solid #1c2029; font-size:.66rem; line-height:1.5; color:#5a6472; text-align:left; }}
  .legal b {{ color:#8a94a3; font-weight:700; }}
</style></head>
<body>
  <form class="card" method="POST" action="/login">
    <div class="brand"><span>nova</span>.byAletheya</div>
    <div class="sub">nova.byaletheya.com</div>
    {msg}
    {nxt}
    <input type="password" name="password" placeholder="Password" autofocus autocomplete="current-password">
    <button type="submit">Enter</button>
    <p class="legal"><b>Confidential &amp; Proprietary.</b> This application and all concepts, designs, workflows, data, and content within are the exclusive property of <b>NovAuto</b> and are intended solely for authorized internal use by its personnel. Unauthorized access, use, copying, reproduction, distribution, or disclosure of any ideas or materials herein is strictly prohibited and may result in legal action.</p>
  </form>
</body></html>"""


def toolbox_login_page(message=""):
    msg = f'<p class="err">{message}</p>' if message else ""
    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0"><title>Aletheya Toolbox</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:#05080d; color:#fff;
    min-height:100vh; display:flex; align-items:center; justify-content:center;
    background-image: radial-gradient(70% 50% at 50% 35%, rgba(155,109,255,.28), rgba(8,16,30,0) 70%); }}
  .card {{ width:340px; max-width:90vw; background:#101319; border:1px solid #232733; border-radius:16px; padding:34px 28px; text-align:center; box-shadow:0 30px 80px rgba(0,0,0,.5); }}
  .brand {{ font-size:1.5rem; font-weight:800; letter-spacing:-.5px; margin-bottom:4px; }}
  .brand span {{ color:#a98bff; }}
  .sub {{ color:#6b7280; font-size:.72rem; letter-spacing:.12em; text-transform:uppercase; margin-bottom:24px; }}
  input {{ width:100%; background:#0a0d12; border:1px solid #2a2f3a; border-radius:9px; padding:13px 15px; color:#fff; font-size:.95rem; outline:none; }}
  input:focus {{ border-color:#a98bff; }}
  button {{ width:100%; margin-top:12px; padding:13px; border:none; border-radius:9px; background:#8a5cf0; color:#fff; font-weight:600; font-size:.95rem; cursor:pointer; }}
  button:hover {{ background:#7a4ce0; }}
  .err {{ color:#ff6b6b; font-size:.82rem; margin-bottom:12px; }}
  .legal {{ margin-top:20px; padding-top:16px; border-top:1px solid #1c2029; font-size:.66rem; line-height:1.5; color:#5a6472; text-align:left; }}
  .legal b {{ color:#8a94a3; font-weight:700; }}
</style></head>
<body>
  <form class="card" method="POST" action="/toolbox-login">
    <div class="brand">Aletheya <span>Toolbox</span></div>
    <div class="sub">Internal tools</div>
    {msg}
    <input type="password" name="password" placeholder="Password" autofocus autocomplete="current-password">
    <button type="submit">Enter</button>
    <p class="legal"><b>Confidential &amp; Proprietary.</b> This workspace and all content within are the exclusive property of <b>Aletheya</b> and intended solely for authorized internal use. Unauthorized access, use, or disclosure is strictly prohibited.</p>
  </form>
</body></html>"""


@app.route("/toolbox-login", methods=["GET", "POST"])
def toolbox_login():
    if request.method == "POST":
        ip = _client_ip() + ":tb"
        cnt, t0 = _LOGIN_FAILS.get(ip, (0, time.time()))
        if time.time() - t0 > _LOGIN_WINDOW:
            cnt, t0 = 0, time.time()
        if cnt >= _LOGIN_MAX:
            return toolbox_login_page("Too many attempts — wait a few minutes and try again."), 429
        if hmac.compare_digest(request.form.get("password", ""), TOOLBOX_PASSWORD):
            session["toolbox_ok"] = True
            _LOGIN_FAILS.pop(ip, None)
            return redirect("/toolbox")
        _LOGIN_FAILS[ip] = (cnt + 1, t0)
        return toolbox_login_page("Incorrect password — try again."), 401
    if session.get("toolbox_ok"):
        return redirect("/toolbox")
    return toolbox_login_page()


def nova_admin_login_page(message="", sel="edgar"):
    msg = f'<p class="err">{message}</p>' if message else ""
    colors = {"edgar": "#3a8eef", "nema": "#f472b6", "arvin": "#8a5cf0"}
    pills = "".join(
        f'<button type="button" class="who{" on" if u["id"]==sel else ""}" data-u="{u["id"]}" onclick="pick(this)">'
        f'<span class="av" style="background:{colors.get(u["id"], "#3a8eef")}">{u["name"][0]}</span>{u["name"]}</button>'
        for u in NOVA_USERS)
    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0"><title>Nova Admins</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:#05080d; color:#fff;
    min-height:100vh; display:flex; align-items:center; justify-content:center;
    background-image: radial-gradient(70% 50% at 50% 35%, rgba(52,211,153,.22), rgba(8,16,30,0) 70%); }}
  .card {{ width:360px; max-width:90vw; background:#101319; border:1px solid #232733; border-radius:16px; padding:32px 28px; text-align:center; box-shadow:0 30px 80px rgba(0,0,0,.5); }}
  .brand {{ font-size:1.5rem; font-weight:800; letter-spacing:-.5px; margin-bottom:4px; }}
  .brand span {{ color:#34d399; }}
  .sub {{ color:#6b7280; font-size:.72rem; letter-spacing:.12em; text-transform:uppercase; margin-bottom:22px; }}
  .who-row {{ display:flex; gap:8px; margin-bottom:16px; }}
  .who {{ flex:1; display:flex; flex-direction:column; align-items:center; gap:7px; padding:12px 4px; border-radius:12px;
    border:1px solid #2a2f3a; background:#0a0d12; color:#8a94a3; font-size:.82rem; font-weight:600; cursor:pointer; transition:.13s; }}
  .who:hover {{ color:#fff; border-color:#3a4150; }}
  .who.on {{ color:#fff; border-color:#34d399; background:rgba(52,211,153,.1); }}
  .who .av {{ width:34px; height:34px; border-radius:50%; display:flex; align-items:center; justify-content:center; font-weight:700; color:#fff; font-size:.95rem; }}
  input {{ width:100%; background:#0a0d12; border:1px solid #2a2f3a; border-radius:9px; padding:13px 15px; color:#fff; font-size:.95rem; outline:none; }}
  input:focus {{ border-color:#34d399; }}
  button.go {{ width:100%; margin-top:12px; padding:13px; border:none; border-radius:9px; background:#1f9d6b; color:#fff; font-weight:600; font-size:.95rem; cursor:pointer; }}
  button.go:hover {{ background:#1b8a5e; }}
  .err {{ color:#ff6b6b; font-size:.82rem; margin-bottom:12px; }}
  .legal {{ margin-top:20px; padding-top:16px; border-top:1px solid #1c2029; font-size:.66rem; line-height:1.5; color:#5a6472; text-align:left; }}
  .legal b {{ color:#8a94a3; font-weight:700; }}
</style></head>
<body>
  <form class="card" method="POST" action="/nova-admins-login">
    <div class="brand">Nova <span>Admins</span></div>
    <div class="sub">Back office · Sign in</div>
    {msg}
    <div class="who-row">{pills}</div>
    <input type="hidden" name="user" id="userField" value="{sel}">
    <input type="password" name="password" placeholder="Your password" autofocus autocomplete="current-password">
    <button class="go" type="submit">Sign in</button>
    <p class="legal"><b>Confidential &amp; Proprietary.</b> Financial, commission, and client records belonging to <b>NovAuto</b>, for authorized administrators only. Unauthorized access, use, or disclosure is strictly prohibited and may result in legal action.</p>
  </form>
  <script>
    function pick(b){{ document.querySelectorAll('.who').forEach(x=>x.classList.remove('on')); b.classList.add('on');
      document.getElementById('userField').value=b.dataset.u; }}
  </script>
</body></html>"""


@app.route("/nova-admins-login", methods=["GET", "POST"])
def nova_admins_login():
    if request.method == "POST":
        ip = _client_ip() + ":na"
        cnt, t0 = _LOGIN_FAILS.get(ip, (0, time.time()))
        if time.time() - t0 > _ADMIN_WINDOW:
            cnt, t0 = 0, time.time()
        if cnt >= _ADMIN_MAX:
            _audit("admin_login_lockout")
            return nova_admin_login_page("Too many attempts — wait a few minutes and try again."), 429
        uid = (request.form.get("user", "") or "").strip().lower()
        pw = request.form.get("password", "")
        roster = _nova_roster()
        rec = roster.get(uid)
        ok = False
        if rec:
            if rec.get("pass") and _check_pw(pw, rec["pass"]):
                ok = True
            elif uid == NOVA_ACCOUNT_ADMIN and NOVA_ADMIN_PASSWORD and hmac.compare_digest(pw, NOVA_ADMIN_PASSWORD):
                ok = True   # Edgar bootstrap / recovery via the env password
        if ok:
            was_site = session.get("ok")           # keep their Nova-workspace login alive
            was_tb = session.get("toolbox_ok")
            session.clear()                        # fresh session id (anti-fixation)
            if was_site:
                session["ok"] = True
            if was_tb:
                session["toolbox_ok"] = True
            session["admin_ok"] = True
            session["nova_user"] = uid
            session["nova_role"] = rec["role"]
            session.permanent = True               # persist the cookie across browser restarts
            session["admin_boot"] = SERVER_BOOT_ID  # tie the session to this server process
            _LOGIN_FAILS.pop(ip, None)
            _audit("admin_login_ok", user=uid)
            return redirect("/nova-admins")
        _LOGIN_FAILS[ip] = (cnt + 1, t0)
        _audit("admin_login_fail", user=uid, attempt=cnt + 1)
        time.sleep(0.3)                            # slow down guessing
        no_pw = bool(rec) and not rec.get("pass") and uid != NOVA_ACCOUNT_ADMIN
        hint = "No password set for that account yet — ask Edgar to set it in Team." if no_pw else "Wrong password — try again."
        return nova_admin_login_page(hint, sel=uid if uid in NOVA_USER_IDS else "edgar"), 401
    if session.get("admin_ok"):
        return redirect("/nova-admins")
    return nova_admin_login_page()


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        ip = _client_ip()
        cnt, t0 = _LOGIN_FAILS.get(ip, (0, time.time()))
        if time.time() - t0 > _LOGIN_WINDOW:        # reset the window
            cnt, t0 = 0, time.time()
        if cnt >= _LOGIN_MAX:                        # too many tries -> back off
            return login_page("Too many attempts — wait a few minutes and try again."), 429
        # constant-time compare avoids leaking the password via timing
        if hmac.compare_digest(request.form.get("password", ""), APP_PASSWORD):
            session.clear()                          # new session id on login (anti-fixation)
            session["ok"] = True
            _LOGIN_FAILS.pop(ip, None)
            nxt = request.args.get("next") or request.form.get("next") or "/novauto"
            if not nxt.startswith("/"):              # only allow local redirects
                nxt = "/novauto"
            return redirect(nxt)
        _LOGIN_FAILS[ip] = (cnt + 1, t0)
        return login_page("Incorrect password — try again.",
                          request.form.get("next", "/novauto")), 401
    return login_page(next_path=request.args.get("next", "/novauto"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


@app.route("/unlock", methods=["GET", "POST"])
def unlock():
    """Second-tier unlock for the restricted tools."""
    if request.method == "POST":
        if hmac.compare_digest((request.json or {}).get("password", ""), RESTRICTED_PASSWORD):
            session["restricted_ok"] = True
            return jsonify({"ok": True})
        return jsonify({"error": "Incorrect password."}), 401
    return jsonify({"unlocked": bool(session.get("restricted_ok"))})


# Paths anyone can reach without logging in: the public byAletheya home page,
# the login screen, the brand asset the home page shows, and public event pages.
PUBLIC_PATHS = ("/", "/login", "/logo.png", "/odyssey")

@app.before_request
def require_login():
    p = request.path
    if p in PUBLIC_PATHS:
        return None
    # The Aletheya Toolbox is its own workspace gated by its own password
    # (session["toolbox_ok"]) — completely independent of the Nova login.
    if p.startswith("/toolbox"):
        if p == "/toolbox-login":
            return None
        if session.get("toolbox_ok"):
            return None
        if request.method == "GET":
            return redirect("/toolbox-login")
        return jsonify({"error": "This area is locked.", "locked": True}), 403
    # Nova Admins is its own workspace gated by its own password (session["admin_ok"]).
    if p.startswith("/nova-admins"):
        if p == "/nova-admins-login":
            return None
        # scripted access: shared-secret header (no Origin requirements — curl/cron)
        tok = request.headers.get("X-Nova-Token", "")
        if NOVA_ADMIN_TOKEN and tok and hmac.compare_digest(tok, NOVA_ADMIN_TOKEN):
            if request.method != "GET":
                _audit("token_write", path=p)
            return None
        if session.get("admin_ok"):
            # Signed in for as long as this server process is up — no idle/absolute time-out.
            # A restart rotates SERVER_BOOT_ID, so a session stamped by an older process is
            # dropped here (the only automatic sign-out). Explicit Lock/logout still works.
            if session.get("admin_boot") != SERVER_BOOT_ID:
                session.pop("admin_ok", None)      # server was refreshed -> re-login
                _audit("admin_session_expired", reason="server_refresh")
            else:
                # cross-site write protection: browser writes must come from us
                if request.method != "GET":
                    origin = request.headers.get("Origin", "")
                    if origin:
                        from urllib.parse import urlsplit
                        if urlsplit(origin).netloc != request.host:
                            _audit("blocked_cross_origin", path=p, origin=origin)
                            return jsonify({"error": "Cross-origin request blocked."}), 403
                return None
        if request.method == "GET":
            return redirect("/nova-admins-login")
        _audit("unauthorized_write", path=p)
        return jsonify({"error": "This area is locked.", "locked": True}), 403
    # Everything else is the Nova workspace, gated by APP_PASSWORD.
    if session.get("ok"):
        return None
    # remember where they were headed so login can send them straight back
    if request.method == "GET":
        return redirect("/login?next=" + request.path)
    return redirect("/login")


@app.before_request
def gate_restricted():
    # block the restricted tools' data endpoints until the second password is entered
    if session.get("restricted_ok"):
        return None
    p = request.path
    if any(p.startswith(pre) for pre in RESTRICTED_API):
        return jsonify({"error": "This area is locked.", "locked": True}), 403


@app.after_request
def security_headers(resp):
    resp.headers["X-Content-Type-Options"] = "nosniff"            # no MIME sniffing
    resp.headers["X-Frame-Options"] = "DENY"                       # no clickjacking/embedding
    resp.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    resp.headers["Permissions-Policy"] = "geolocation=(), camera=()"   # mic allowed: voice dictation
    # browsers remember to only ever use HTTPS for this host
    resp.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    if request.path.startswith("/nova-admins"):
        # financial data: never cached to disk, and the page can only talk to itself —
        # no external scripts, no external requests (blocks exfiltration + injected code)
        resp.headers["Cache-Control"] = "no-store, private"
        resp.headers["Content-Security-Policy"] = (
            "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; "
            "img-src 'self' data:; connect-src 'self'; font-src 'self'; object-src 'none'; "
            "frame-ancestors 'none'; base-uri 'none'; form-action 'self'")
    return resp

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Saved cars live here. Set GENERATED_DIR to a persistent disk/volume path on the
# host so the gallery survives restarts & redeploys (the default app dir is ephemeral).
GENERATED_DIR = os.environ.get("GENERATED_DIR", "").strip() or os.path.join(BASE_DIR, "generated")
os.makedirs(GENERATED_DIR, exist_ok=True)


def load_api_key():
    # Prefer an environment variable (used in deployment). Accept the common names so
    # the key that's ALREADY configured on the host is found regardless of what it was
    # called (Railway setups have used API_KEY as well as OPENAI_API_KEY).
    for name in ("OPENAI_API_KEY", "API_KEY", "OPENAI_KEY", "OPEN_AI_API_KEY"):
        key = os.environ.get(name, "").strip()
        if key:
            if name != "OPENAI_API_KEY":
                print(f"[ai] using the OpenAI key from ${name}")
            return key
    env_path = os.path.join(BASE_DIR, ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                m = re.match(r"\s*(?:OPENAI_API_KEY|API_KEY|OPENAI_KEY)\s*=\s*(.+)", line)
                if m:
                    return m.group(1).strip()
    print("[ai] no OpenAI key found — AI features (Ask Nova, parsers) are off until "
          "OPENAI_API_KEY is set in the host env.")
    return ""


API_KEY = load_api_key()

# One reused OpenAI client instead of constructing a fresh one on every request.
_OAI_CLIENT = None


def _oai():
    global _OAI_CLIENT
    if _OAI_CLIENT is None:
        _OAI_CLIENT = OpenAI(api_key=API_KEY)
    return _OAI_CLIENT


# the catalog (image source). Set CARSXE_API_KEY in the host env.
CARSXE_API_KEY = os.environ.get("CARSXE_API_KEY", "").strip()
_CARSXE_ALLOWED = set()     # exact image URLs the catalog returned (fast-path proxy allow)
_CARSXE_HOSTS = set()       # CDN hosts the catalog served from — learned so a URL still proxies
                            # after the exact-set is lost (server restart) or evicted
_CARSXE_CDN_ROOTS = ('carsxe.com', 'imagin.studio')   # seeded vendor CDNs so proxy works even before the first fetch post-restart
_CARSXE_IMG_CACHE = {}      # vehicle -> image response, so repeats never re-hit the catalog
_CARSXE_META_CACHE = {}     # vehicle -> colors + trims
_CARSXE_CANON_CACHE = {}    # vehicle -> canonical make/model/trim (name cleanup)
_CARSXE_PROXY_CACHE = {}    # image url -> data URL
_CARSXE_CACHE_MAX = 150     # keep the last ~150 of each (FIFO)


def _cache_put(cache, key, val):
    if key in cache:
        return
    if len(cache) >= _CARSXE_CACHE_MAX:
        cache.pop(next(iter(cache)))   # evict oldest
    cache[key] = val


def _carsxe_base_model(model):
    """Drop the powertrain word the model now carries for display (e.g. "Sorento Hybrid"
    -> "Sorento") so the image/colour catalog — which only knows the base model — still
    matches. Whole-word only, so "Mach-E" / "EV6" / "bZ4X" survive untouched."""
    q = re.sub(r"\b(?:plug[\s-]?in hybrid|plug[\s-]?in|hybrid|phev|electric|ev|bev)\b", " ", model or "", flags=re.I)
    q = re.sub(r"\s+", " ", q).strip()
    return q or (model or "")


def _carsxe_host(url):
    try:
        return urllib.parse.urlparse(url).netloc.lower().split(":")[0]
    except Exception:
        return ""


def _carsxe_remember(url):
    """Record a catalog image URL (and its CDN host) as safe to proxy."""
    if not url:
        return
    _CARSXE_ALLOWED.add(url)
    h = _carsxe_host(url)
    if h:
        _CARSXE_HOSTS.add(h)


def _carsxe_proxy_ok(url):
    """Proxy guard: exact URL we've seen, OR any URL on a CDN host the catalog has
    served from / a known vendor CDN. Host-based so a thumbnail still loads after the
    exact allow-set is lost to a restart or FIFO eviction (was a silent 'URL not allowed')."""
    if url in _CARSXE_ALLOWED:
        return True
    h = _carsxe_host(url)
    if not h:
        return False
    if h in _CARSXE_HOSTS:
        return True
    return any(h == r or h.endswith("." + r) for r in _CARSXE_CDN_ROOTS)


# CarVector (trial image source). Set CARVECTOR_API_KEY in the host env.
CARVECTOR_API_KEY = os.environ.get("CARVECTOR_API_KEY", "").strip()
_CARVECTOR_ALLOWED = set()
_CARVECTOR_CACHE = {}

# Art-directed render rules: locked framing/scale/lighting for a consistent campaign look,
# rendered on a TRANSPARENT background so each car drops cleanly onto the ad template.
RENDER_RULES = (
    "MANDATORY CONSISTENCY RULES — NEVER CHANGE:\n"
    "• Vehicle centered in frame • Front 3/4 angle • Facing slightly left • Entire vehicle "
    "visible with identical spacing around the car in every image • Landscape 3:2 wide aspect ratio "
    "• Fully transparent background (PNG alpha) — no background fill, no backdrop, no scenery "
    "• No floor • No shadows • No reflections • No gradients • No environment "
    "• Car windows fully opaque dark tinted black glass • Cool cinematic showroom lighting "
    "• Photorealistic ultra-high-resolution rendering • Sharp edges with clean cutout separation "
    "from the transparent background\n"
    "CRITICAL SCALE & FRAMING LOCK:\n"
    "The vehicle MUST occupy the EXACT SAME percentage of the frame in every image.\n"
    "DO NOT: • zoom in or out • change focal length • change camera distance • alter crop "
    "• alter perspective • alter wheel positioning • alter roof height within frame • alter "
    "tire-to-bottom spacing • alter spacing above vehicle • alter visual weight of vehicle in canvas\n"
    "The wheels, roofline, and body proportions must align consistently across all generated "
    "vehicles so the entire set looks like one professionally art-directed automotive campaign.\n"
    "Keep identical camera height, lens perspective, framing, vehicle scale, crop margins, "
    "lighting direction, and angle for every vehicle.\n"
    "Render the image only — no text, watermarks, or captions anywhere in the image."
)


def research_vehicle(client, vehicle, bodystyle):
    """Look up the real, current design of this exact model year on the web so the
    image model renders THIS year's car (not an outdated training-data version).
    Best-effort: returns "" on any failure so generation still proceeds."""
    query = (
        f"I need an accurate visual reference for the {vehicle}"
        + (f" ({bodystyle})" if bodystyle else "")
        + ".\n"
        "STEP 1: Search the web and determine which GENERATION of this model is sold BRAND-NEW for "
        "that exact model year. Identify the generation name / chassis code and the year that "
        "generation (or its facelift) was introduced. Many models were recently redesigned, and older "
        "generations dominate image search — you MUST use the CURRENT/NEWEST generation for this model "
        "year and IGNORE every previous generation, even if older photos are more common.\n"
        "STEP 2: In 130-170 words, describe ONLY the exterior of THAT current generation so an "
        "illustrator can draw it accurately. Start by stating the generation code and its years in "
        "production, then describe: front grille shape and pattern, headlight and daytime-running-light "
        "signature, front bumper and air intakes, overall body proportions and roofline, side character "
        "lines, wheel design, and badge placement. Explicitly call out what makes THIS generation look "
        "different from the previous one. Be factual and specific — no pricing, no interior, no fluff."
    )
    last_err = None
    for tool_type in ("web_search", "web_search_preview"):
        try:
            resp = client.responses.create(
                model="gpt-4.1",
                tools=[{"type": tool_type}],
                input=query,
            )
            text = (getattr(resp, "output_text", "") or "").strip()
            if text:
                return text
        except Exception as e:
            last_err = e
            continue
    if last_err:
        print(f"[research] skipped ({last_err})")
    return ""


def classify_error(e):
    """Turn an OpenAI/SDK exception into a clear, human reason + the raw detail.
    `fatal` means there's no point continuing a bulk run (key/quota/access issues)."""
    name = type(e).__name__
    status = getattr(e, "status_code", None)
    code = getattr(e, "code", None)
    detail = str(getattr(e, "message", "") or e)
    low = f"{code} {detail}".lower()

    if "insufficient_quota" in low or "exceeded your current quota" in low or "billing" in low:
        reason, fatal = ("You're out of OpenAI credits/quota — add billing or credits to the "
                         "OpenAI account.", True)
    elif status == 401 or name == "AuthenticationError" or "api key" in low:
        reason, fatal = ("Invalid or missing OpenAI API key on the server.", True)
    elif status == 403 or name == "PermissionDeniedError" or ("verified" in low and "organization" in low):
        reason, fatal = ("This OpenAI account can't use the image model yet (gpt-image-1 may "
                         "require organization verification).", True)
    elif status == 429 or name == "RateLimitError":
        reason, fatal = ("Hit the OpenAI rate limit — wait a moment and try again.", False)
    elif status == 400 or name == "BadRequestError":
        reason, fatal = ("Request was rejected (possibly content policy or an invalid prompt).", False)
    elif name in ("APIConnectionError", "APITimeoutError"):
        reason, fatal = ("Couldn't reach OpenAI (network/timeout) — check the connection and retry.", False)
    elif status and status >= 500:
        reason, fatal = ("OpenAI had a server error — try again shortly.", False)
    else:
        reason, fatal = ("Image generation failed.", False)

    return {"reason": reason, "detail": detail, "fatal": fatal,
            "code": code, "status": status or 500}


def build_image_prompt(vehicle, bodystyle, color, reference):
    head = f"Render a single studio product photo of this exact vehicle: {vehicle}.\n"
    if bodystyle:
        head += (f"Body style: {bodystyle} — match the silhouette, roofline, doors, and "
                 f"proportions to this body style.\n")
    head += f"Exterior paint color: {color or 'the standard factory color for this model'}.\n"
    if reference:
        head += ("\nAUTHORITATIVE REAL-WORLD REFERENCE (from current sources). Render the CURRENT/NEWEST "
                 "generation of this model for this model year EXACTLY as described — never an older "
                 f"generation, even if older designs are more familiar:\n{reference}\n")
    return head + "\n" + RENDER_RULES


def car_slug(vehicle, bodystyle, color):
    """Deterministic identity for a car image: same vehicle+body+color -> same file.
    Lowercased so matching is case-insensitive across platforms."""
    key = " ".join(p for p in (vehicle, bodystyle, color) if p).lower()
    return re.sub(r'[^\w\s-]', '', key).strip().replace(' ', '_')[:90]


# The public byAletheya home page — the front door at byaletheya.com. Anyone can
# see it; the "Open Novauto" button links to /novauto, which is behind the login.
@app.route("/")
def home():
    return send_file(os.path.join(BASE_DIR, "home.html"))


# Public event landing page: 2000 — A Rave Odyssey (Jubiland, 21+).
@app.route("/odyssey")
def odyssey():
    return send_file(os.path.join(BASE_DIR, "odyssey.html"))


# Each tool has its own clean URL. They all serve the same single-page app; the
# client reads the path on load to open the right view (and pushState keeps the
# URL in sync as you switch tools). Deep-linkable and bookmarkable.
# /novauto is the entry point reached from the home page.
@app.route("/novauto")
@app.route("/lease")
@app.route("/sold")
@app.route("/leasead-test")
@app.route("/deal-hub")
@app.route("/review-generator")
@app.route("/invoice")
@app.route("/quote")
@app.route("/desking")
def index():
    return send_file(os.path.join(BASE_DIR, "index.html"))


@app.route("/logo.png")
def logo():
    return send_file(os.path.join(BASE_DIR, "logo.png"))


@app.route("/background.png")
def background():
    return send_file(os.path.join(BASE_DIR, "background.png"))


@app.route("/sold-bg.png")
def sold_bg():
    return send_file(os.path.join(BASE_DIR, "sold-bg.png"))


@app.route("/bg-economy.png")
def bg_economy():
    return send_file(os.path.join(BASE_DIR, "bg-economy.png"))


@app.route("/nova-plate.png")
def nova_plate():
    return send_file(os.path.join(BASE_DIR, "nova-plate.png"))


def _nova_public_store():
    """The store as the browser may see it — password hashes stripped out."""
    store = _nova_load()
    if isinstance(store.get("users"), list):
        store = dict(store)
        store["users"] = [{k: v for k, v in u.items() if k != "pass"} for u in store["users"]]
    return store


def _nova_admin_serve(filename):
    """Serve a Nova Admins tool page with the shared dataset + the signed-in user
    injected (password hashes stripped; the gitignored store stays out of git)."""
    html = open(os.path.join(BASE_DIR, filename), encoding="utf-8").read()
    seed = json.dumps(_nova_public_store())
    user = _nova_current_user() or {}
    inject = "window.NOVA_SEED=" + seed + ";window.NOVA_USER=" + json.dumps(user) + ";"
    return html.replace("/*NOVA_SEED*/", inject, 1)


# Nova Admins is a SUITE: each tool is its own page under /nova-admins/*
# (all behind the same admin gate + sharing one data store).
@app.route("/nova-admins")
def nova_admins_page():
    """Tool 1: the deal ledger / agent payroll."""
    return _nova_admin_serve("nova_admins.html")


@app.route("/nova-admins/tasks")
def nova_admins_tasks_page():
    """Tool 2: the Linear-style task manager."""
    return _nova_admin_serve("nova_tasks.html")


@app.route("/nova-admins/notes")
def nova_admins_notes_page():
    """Tool 3: Notion-style notes."""
    return _nova_admin_serve("nova_notes.html")


@app.route("/nova-admins/parse-task", methods=["POST"])
def nova_admins_parse_task():
    """Turn natural language into one or more structured tasks (title, subtasks,
    assignee, due date, priority, labels, notes)."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    body = request.json or {}
    text = (body.get("text") or "").strip()
    if not text:
        return jsonify({"error": "Say what needs doing."}), 400
    today = (body.get("today") or "").strip()
    task_schema = {
        "type": "object", "additionalProperties": False,
        "required": ["title", "status", "priority", "assignees", "due", "labels", "notes", "subtasks"],
        "properties": {
            "title": {"type": "string"},
            "status": {"type": "string", "enum": ["backlog", "todo", "inprogress", "done", ""]},
            "priority": {"type": "string", "enum": ["urgent", "high", "medium", "low", "none", ""]},
            "assignees": {"type": "array", "items": {"type": "string", "enum": ["nema", "arvin", "edgar"]}},
            "due": {"type": "string"},
            "labels": {"type": "array", "items": {"type": "string"}},
            "notes": {"type": "string"},
            "subtasks": {"type": "array", "items": {"type": "string"}},
        },
    }
    schema = {
        "type": "object", "additionalProperties": False,
        "required": ["ok", "summary", "tasks"],
        "properties": {"ok": {"type": "boolean"}, "summary": {"type": "string"},
                       "tasks": {"type": "array", "items": task_schema}},
    }
    sys = (
        "You convert natural language into structured to-do tasks for Nova's back-office task board. "
        "Output STRICT JSON. Rules:\n"
        f"- Today is {today or 'unknown'} — resolve relative dates ('tomorrow', 'Friday', 'end of month', 'in 2 weeks') "
        "to YYYY-MM-DD in `due`; '' if no date mentioned.\n"
        "- One message may contain SEVERAL tasks — split them; but an enumeration of steps under one goal is ONE task "
        "with `subtasks` (short imperative strings).\n"
        "- title: short imperative ('Pay agents for June'). Put extra context/details in `notes`.\n"
        "- assignees: array of nema / arvin / edgar for EVERYONE the task is for ('me' = edgar); [] if nobody named. "
        "Multiple people allowed — 'Nema and Edgar' -> [\"nema\",\"edgar\"].\n"
        "- priority: only if urgency is expressed ('asap'/'urgent' -> urgent; 'important' -> high; 'whenever/low prio' -> low), else ''.\n"
        "- status: 'todo' unless they say it's already started ('inprogress'), an idea/someday ('backlog'), or done.\n"
        "- labels: 1-2 short category tags ONLY if obvious (Payroll, Collections, Deals, Follow-up, Ops, Automation, Finance).\n"
        "- ok=false only if the message clearly isn't a task. summary = one short confirmation sentence."
    )
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "system", "content": sys}, {"role": "user", "content": text}],
            response_format={"type": "json_schema", "json_schema": {"name": "tasks", "strict": True, "schema": schema}},
        )
        return jsonify(json.loads(resp.choices[0].message.content))
    except Exception as e:
        return jsonify({"error": "Couldn't parse that — " + str(e)[:160]}), 500


@app.route("/nova-admins/parse", methods=["POST"])
def nova_admins_parse():
    """Turn a natural-language / pasted deal into one structured Nova Admins deal."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    body = request.json or {}
    text = (body.get("text") or "").strip()
    if not text:
        return jsonify({"error": "Type a deal to add."}), 400
    agents = [a for a in (body.get("agents") or []) if isinstance(a, dict) and a.get("id")][:60]
    today = (body.get("today") or "").strip()
    agent_list = "; ".join(f'{a.get("id")}={a.get("name")}' for a in agents) or "(none provided)"
    schema = {
        "type": "object", "additionalProperties": False,
        "required": ["client", "year", "make", "model", "vin", "dealer", "type", "term", "agentId",
                     "lead", "source", "front", "back", "feeReferral", "feeProgram", "pay", "payBack", "wOffered", "wSold",
                     "fColl", "bColl", "aPaidF", "aPaidFd", "aPaidB", "aPaidBd", "date", "notes", "ok", "summary"],
        "properties": {
            "client": {"type": "string"}, "year": {"type": "string"}, "make": {"type": "string"},
            "model": {"type": "string"}, "vin": {"type": "string"}, "dealer": {"type": "string"},
            "type": {"type": "string", "enum": ["Lease", "Buy", ""]}, "term": {"type": "string"},
            "agentId": {"type": "string"}, "lead": {"type": "string", "enum": ["own", "nova", "referral", ""]},
            "source": {"type": "string", "enum": ["FB", "IG", "Google", "Yelp", "Referral", "Repeat", "Walk-in", "Website", "Other", ""]},
            "front": {"type": "number"}, "back": {"type": "number"},
            "feeReferral": {"type": "number"}, "feeProgram": {"type": "number"},
            "pay": {"type": "string", "enum": ["Stripe", "Zelle", "Cash", "Check", ""]},
            "payBack": {"type": "string", "enum": ["Zelle", "Check", "ACH", "Wire", "Cash", ""]},
            "wOffered": {"type": "boolean"}, "wSold": {"type": "boolean"},
            "fColl": {"type": "boolean"}, "bColl": {"type": "boolean"},
            "aPaidF": {"type": "boolean"}, "aPaidFd": {"type": "string"},
            "aPaidB": {"type": "boolean"}, "aPaidBd": {"type": "string"},
            "date": {"type": "string"}, "notes": {"type": "string"},
            "ok": {"type": "boolean"}, "summary": {"type": "string"},
        },
    }
    sys = (
        "You convert ONE natural-language or pasted car lease/finance deal into a structured deal for "
        "Nova's back-office ledger. Output STRICT JSON for the schema. Rules:\n"
        f"- Today is {today or 'unknown'}; dates are YYYY-MM-DD. If no deal date is given, use today.\n"
        "- client: customer name as 'First L.' (first name + last initial).\n"
        f"- agentId: match the salesperson to ONE of these agents and output its id (or '' if unknown): {agent_list}\n"
        "- lead: 'own' = agent's own lead/Agent-sourced; 'nova' = Nova Lead (Nova paid for it); 'referral' = referral. This sets the agent split.\n"
        "- source = acquisition CHANNEL for metrics (separate from lead/split): 'FB' or 'IG' (Meta ads), 'Google', 'Yelp', "
        "'Referral', 'Repeat', 'Walk-in', 'Website', 'Other'; '' if unknown. A Nova lead is usually FB or IG.\n"
        "- type: Lease or Buy. term = months (string).\n"
        "- front = front gross (client/broker fee), back = back gross (dealer reserve). Numbers only — no $ or commas.\n"
        "- feeProgram = a program fee if stated. NEVER include an Envy fee anywhere — Envy (20% of back) is computed "
        "automatically by the ledger.\n"
        "- feeProgram = a program fee OR a fee shared with Jason (Jason is part of Program now). feeReferral = a referral/shared fee OR a generic 'fees shared' "
        "lump. Do NOT include Stripe processing (its 3% is auto-computed from the payment method).\n"
        "- pay = how the FRONT was charged (Stripe/Zelle/Cash/Check). Stripe is a FRONT-END client card charge and its 3% "
        "is auto-computed. payBack = how the BACK (dealer paying Nova) was paid (Check/ACH/Wire/Zelle/Cash), '' if no back end. "
        "The back end is NEVER Stripe.\n"
        "- wOffered/wSold = was warranty offered / sold.\n"
        "- fColl = was the front (client) money collected by Nova; aPaidF = was the agent paid their front share "
        "(aPaidFd = that pay date if stated).\n"
        "- BACK-SIDE money is NEVER inferred: always output bColl=false and aPaidB=false (aPaidBd='') no matter "
        "what the text says — Edgar confirms back-end money himself in the ledger.\n"
        "- Unknown strings => '', unknown numbers => 0, unknown booleans => false.\n"
        "- ok=true if you found at least a client or a vehicle; false if this isn't a deal. "
        "summary = one short sentence describing what you logged."
    )
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "system", "content": sys}, {"role": "user", "content": text}],
            response_format={"type": "json_schema", "json_schema": {"name": "deal", "strict": True, "schema": schema}},
        )
        out = json.loads(resp.choices[0].message.content)
        # Hard rule (Edgar): back-side money is only ever marked by him, once he confirms
        # receipt — never inferred from pasted text, whatever the model returned.
        out["bColl"] = False
        out["aPaidB"] = False
        out["aPaidBd"] = ""
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": "Couldn't parse that — " + str(e)[:160]}), 500


NOVA_ADMIN_DB = lambda: os.path.join(GENERATED_DIR, "nova_admins.json")


def _nova_nightly_backup():
    """Daily dated snapshot of the Nova Admins data → generated/nova_admins_backups/,
    keeping the last 30 days. Runs one snapshot per calendar day, survives restarts."""
    import shutil
    while True:
        try:
            src = NOVA_ADMIN_DB()
            if os.path.exists(src):
                bdir = os.path.join(GENERATED_DIR, "nova_admins_backups")
                os.makedirs(bdir, exist_ok=True)
                dest = os.path.join(bdir, "nova_admins_" + time.strftime("%Y-%m-%d") + ".json")
                if not os.path.exists(dest):
                    shutil.copy2(src, dest)
                for old in sorted(f for f in os.listdir(bdir) if f.endswith(".json"))[:-30]:
                    try:
                        os.remove(os.path.join(bdir, old))
                    except Exception:
                        pass
        except Exception:
            pass
        time.sleep(6 * 3600)   # re-check every 6h → at most one snapshot per day


try:
    threading.Thread(target=_nova_nightly_backup, daemon=True).start()
except Exception:
    pass


# Serialize all writes to the shared store so concurrent edits can't interleave.
_NOVA_LOCK = threading.Lock()


def _nova_load():
    path = NOVA_ADMIN_DB()
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"agents": [], "deals": [], "expenses": []}


def _nova_write(data):
    path = NOVA_ADMIN_DB()
    if os.path.exists(path):
        try:
            import shutil
            shutil.copy2(path, path + ".bak")
        except Exception:
            pass
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=1)
    os.replace(tmp, path)


@app.route("/nova-admins/data", methods=["GET"])
def nova_admins_data():
    """Serve the current shared dataset for the live tool (password hashes stripped)."""
    return jsonify(_nova_public_store())


@app.route("/nova-admins/whoami", methods=["GET"])
def nova_admins_whoami():
    return jsonify(_nova_current_user() or {})


@app.route("/nova-admins/users", methods=["GET"])
def nova_admins_users():
    """Team roster (names/roles/whether a password is set) — never the hashes."""
    roster = _nova_roster()
    return jsonify({
        "users": [{"id": u["id"], "name": u["name"], "role": u["role"], "hasPass": bool(u.get("pass"))}
                  for u in roster.values()],
        "me": _nova_current_user(), "accountAdmin": NOVA_ACCOUNT_ADMIN,
    })


@app.route("/nova-admins/set-password", methods=["POST"])
def nova_admins_set_password():
    me = _nova_current_user()
    if not me:
        return jsonify({"error": "Not signed in."}), 403
    body = request.json or {}
    target = (body.get("user") or me["id"]).strip().lower()
    pw = body.get("password") or ""
    if len(pw) < 6:
        return jsonify({"error": "Password must be at least 6 characters."}), 400
    if target != me["id"] and me["id"] != NOVA_ACCOUNT_ADMIN:
        return jsonify({"error": "Only Edgar can set another person's password."}), 403
    if not _nova_set_password(target, pw):
        return jsonify({"error": "Unknown user."}), 400
    _audit("set_password", by=me["id"], target=target)
    return jsonify({"ok": True})


@app.route("/nova-admins/logout")
def nova_admins_logout():
    for k in ("admin_ok", "nova_user", "nova_role", "admin_at", "admin_seen", "admin_boot"):
        session.pop(k, None)
    _audit("admin_logout")
    return redirect("/nova-admins-login")


@app.route("/nova-admins/save", methods=["POST"])
def nova_admins_save():
    """Replace the WHOLE dataset (used by Import / seeding). Full overwrite by design."""
    data = request.get_json(silent=True) or {}
    if not all(isinstance(data.get(k), list) for k in ("agents", "deals", "expenses")):
        return jsonify({"error": "Expected agents, deals and expenses arrays."}), 400
    if len(json.dumps(data)) > 12_000_000:
        return jsonify({"error": "Payload too large."}), 413
    try:
        with _NOVA_LOCK:
            # tasks/notes/users survive an Import from an older export that lacks them.
            # users (accounts + password hashes) are NEVER taken from the payload.
            existing = _nova_load()
            clean = {"agents": data["agents"], "deals": data["deals"], "expenses": data["expenses"],
                     "tasks": data.get("tasks", existing.get("tasks", [])),
                     "notes": data.get("notes", existing.get("notes", [])),
                     "users": existing.get("users", [])}
            _nova_write(clean)
        _audit("data_replace", deals=len(clean["deals"]), expenses=len(clean["expenses"]),
               tasks=len(clean["tasks"]), notes=len(clean["notes"]))
        return jsonify({"ok": True, "deals": len(clean["deals"]), "expenses": len(clean["expenses"])})
    except Exception as e:
        return jsonify({"error": str(e)[:160]}), 500


@app.route("/nova-admins/mutate", methods=["POST"])
def nova_admins_mutate():
    """Row-level write: upsert or delete ONE deal/agent/expense by id, merged into the
    shared store under a lock. Concurrent edits to different rows never clobber."""
    body = request.get_json(silent=True) or {}
    try:
        with _NOVA_LOCK:
            data = _nova_load()
            for coll, key in (("deals", "deal"), ("agents", "agent"), ("expenses", "expense"), ("tasks", "task"), ("notes", "note")):
                item = body.get(key)
                if isinstance(item, dict):
                    arr = data.setdefault(coll, [])
                    for i, x in enumerate(arr):
                        if x.get("id") == item.get("id"):
                            arr[i] = item
                            break
                    else:
                        arr.append(item)
                delk = "delete" + key[0].upper() + key[1:]
                if delk in body:
                    data[coll] = [x for x in data.get(coll, []) if x.get("id") != body[delk]]
            _nova_write(data)
            changed = [k for k in ("deal", "agent", "expense", "task", "note") if k in body] \
                + [k for k in body if k.startswith("delete")]
            _audit("mutate", what=",".join(changed) or "none")
            return jsonify({"ok": True, "deals": len(data.get("deals", []))})
    except Exception as e:
        return jsonify({"error": str(e)[:160]}), 500


# ---------------------------------------------------------------------------
# Nova Admins · SYSTEM-WIDE AGENT
# One assistant, stickied on every Admin page. It can ANSWER from live data,
# NAVIGATE anywhere, CREATE deals/tasks/notes, and MODIFY existing records.
# Answers + navigation come back immediately; writes come back as a PROPOSAL
# the user confirms, then /agent/apply executes them under the lock + audit.
# ---------------------------------------------------------------------------

def _n_num(x):
    try:
        return float(x or 0)
    except Exception:
        return 0.0


def _nova_fee_stripe(d):
    """Mirror the client feeStripe(): 3% of the FRONT when it was charged via Stripe.
    Stripe is a front-end (client card) charge only — the back end is the dealer
    paying Nova (Check/ACH/Wire/Zelle), never Stripe. A stored total > 0 overrides."""
    v = d.get("feeStripe")
    if v not in (None, "") and _n_num(v) > 0:
        return _n_num(v)
    return round(_n_num(d.get("front")) * 0.03) if d.get("pay") == "Stripe" else 0.0


_NOVA_ENVY_START = "2026-07-01"   # envy applies to deals dated here onward (Edgar, July 12 2026)


def _nova_fee_envy(d):
    """Mirror the client feeEnvy(): auto 20% of back for every deal dated
    _NOVA_ENVY_START or later (or carrying the field); a stored number is a
    manual override; deals before the start date stay 0."""
    has = "feeEnvy" in d
    v = d.get("feeEnvy")
    if has and v not in (None, ""):
        return _n_num(v)
    if not has and str(d.get("date") or "") < _NOVA_ENVY_START:
        return 0.0
    back = _n_num(d.get("back"))
    return round(back * 0.20) if back > 0 else 0.0


def _nova_calc(d, amap):
    """Mirror the client calc(): netPool = front+back − fees; agentCut = netPool×pct
    (or a flat override); novaCut = netPool − agentCut. Kept in lockstep with nova_admins.html."""
    front, back = _n_num(d.get("front")), _n_num(d.get("back"))
    # Envy is NOT a shared fee — it's a Nova-only cost (Envy keeps 20% of the back),
    # subtracted from Nova below so the agent's split is untouched.
    # feeJason is legacy (Jason folded into Program) — still summed so old, un-re-saved
    # deals reconcile; new deals carry it in feeProgram.
    fees = (_n_num(d.get("feeJason")) + _nova_fee_stripe(d) + _n_num(d.get("feeReferral"))
            + _n_num(d.get("feeProgram")))
    combined = front + back
    net = combined - fees
    pct = ((amap.get(d.get("agentId")) or {}).get("pct") or {}).get(d.get("lead")) or 0
    ov = d.get("override")
    agent = _n_num(ov) if ov not in (None, "") else net * pct / 100.0
    envy = _nova_fee_envy(d)
    # Per-side nets (mirrors client calc): EVERY shared fee — Stripe, Referral, and
    # Program (incl. legacy Jason, e.g. Jason Ma) — comes entirely off the FRONT, never
    # the back. The back end is the dealer paying Nova; these fees are all deducted from
    # the client's front payment. Only when there is no front do the fees fall to the back.
    fees_f = fees if front > 0 else 0.0   # all shared fees ride the front (unless there's no front)
    front_net = front - fees_f if combined else 0.0
    back_net = back - (fees - fees_f) if combined else 0.0
    ratio = agent / (net or 1)
    # Envy is a NOVA-ONLY cost: Envy receives the back, keeps 20% (Nova eats it), forwards
    # 80% to Nova. Subtracted from Nova's take; the agent's split is unaffected.
    return {"fees": fees, "net": net, "agent": agent, "nova": (net - agent) - envy, "envy": envy, "pct": pct,
            "agentFront": front_net * ratio, "agentBack": back_net * ratio}


def _nova_snapshot(store, today, query=""):
    """A compact, grounded picture of the business for the agent: headline metrics,
    agents, the most relevant deals (recent + any matching the query), tasks and notes."""
    agents = store.get("agents", []) or []
    amap = {a.get("id"): a for a in agents}
    deals = store.get("deals", []) or []
    tasks = store.get("tasks", []) or []
    notes = store.get("notes", []) or []
    ym, yr = today[:7], today[:4]
    mtd = ytd = 0.0
    owed_in = owed_out = 0
    to_collect = owed_agents = 0.0
    per_agent = {}
    for d in deals:
        c = _nova_calc(d, amap)
        dt = str(d.get("date") or "")
        if dt[:4] == yr:
            ytd += c["nova"]
        if dt[:7] == ym:
            mtd += c["nova"]
            aid = d.get("agentId") or ""
            pa = per_agent.setdefault(aid, {"name": (amap.get(aid) or {}).get("name", aid or "—"), "cut": 0.0})
            pa["cut"] += c["agent"]
        # money IN owed to Nova (per uncollected side) — matches the dashboard exactly
        f = False
        if _n_num(d.get("front")) > 0 and not d.get("fColl"):
            to_collect += _n_num(d.get("front")); f = True
        if _n_num(d.get("back")) > 0 and not d.get("bColl"):
            to_collect += _n_num(d.get("back")); f = True
        if f:
            owed_in += 1
        # money OUT owed to the agent (per unpaid side with a real agent share)
        p = False
        if c["agentFront"] > 1 and not d.get("aPaidF"):
            owed_agents += c["agentFront"]; p = True
        if c["agentBack"] > 1 and not d.get("aPaidB"):
            owed_agents += c["agentBack"]; p = True
        if p:
            owed_out += 1
    q = [w for w in (query or "").lower().split() if len(w) > 2]

    # expenses: MTD/YTD spend + the recent (or query-matched) line items, with ids so
    # the agent can update/delete a specific one.
    expenses = store.get("expenses", []) or []
    exp_mtd = sum(_n_num(e.get("amt")) for e in expenses if str(e.get("date") or "")[:7] == ym)
    exp_ytd = sum(_n_num(e.get("amt")) for e in expenses if str(e.get("date") or "")[:4] == yr)
    e_ranked = sorted(expenses, key=lambda e: str(e.get("date") or ""), reverse=True)
    e_hit = [e for e in e_ranked if q and any(w in (str(e.get("cat", "")) + " " + str(e.get("desc", ""))).lower() for w in q)][:20]
    e_seen = {id(e) for e in e_hit}
    e_picked = e_hit + [e for e in e_ranked if id(e) not in e_seen][:15]

    def match(d):
        blob = " ".join(str(d.get(k, "")) for k in ("client", "make", "model", "dealer", "agentId", "notes")).lower()
        return any(w in blob for w in q)

    ranked = sorted(deals, key=lambda d: str(d.get("date") or ""), reverse=True)
    picked = [d for d in ranked if q and match(d)][:40]
    seen = {id(d) for d in picked}
    for d in ranked:
        if len(picked) >= 60:
            break
        if id(d) not in seen:
            picked.append(d)

    def deal_row(d):
        c = _nova_calc(d, amap)
        return {"id": d.get("id"), "date": d.get("date"), "client": d.get("client"),
                "vehicle": " ".join(str(d.get(k, "")) for k in ("year", "make", "model")).strip(),
                "agent": (amap.get(d.get("agentId")) or {}).get("name", d.get("agentId")),
                "lead": d.get("lead"), "front": d.get("front"), "back": d.get("back"), "pay": d.get("pay"),
                "collected": {"front": bool(d.get("fColl")), "back": bool(d.get("bColl"))},
                "agentPaid": {"front": bool(d.get("aPaidF")), "back": bool(d.get("aPaidB"))},
                "netPool": round(c["net"]), "agentCut": round(c["agent"]), "novaCut": round(c["nova"])}

    return {
        "today": today,
        "metrics": {"deals_total": len(deals), "mtd_nova_profit": round(mtd), "ytd_nova_profit": round(ytd),
                    "mtd_expenses": round(exp_mtd), "ytd_expenses": round(exp_ytd),
                    "to_collect": round(to_collect), "deals_with_uncollected_money": owed_in,
                    "owed_to_agents": round(owed_agents), "deals_agent_unpaid": owed_out,
                    "per_agent_mtd": [{"name": v["name"], "cut": round(v["cut"])}
                                      for v in sorted(per_agent.values(), key=lambda x: -x["cut"])]},
        "agents": [{"id": a.get("id"), "name": a.get("name"), "pct": a.get("pct")} for a in agents],
        "deals_shown": len(picked), "deals_total": len(deals),
        "deals": [deal_row(d) for d in picked],
        "tasks": [{"id": t.get("id"), "title": t.get("title"), "status": t.get("status"),
                   "priority": t.get("priority"),
                   "assignees": t.get("assignees") or ([t["assignee"]] if t.get("assignee") else []),
                   "due": t.get("due"), "subtasks": len(t.get("subtasks", []))} for t in tasks],
        "notes": [{"id": n.get("id"), "title": n.get("title")} for n in notes],
        "expenses_shown": len(e_picked), "expenses_total": len(expenses),
        "expenses": [{"id": e.get("id"), "date": e.get("date"), "cat": e.get("cat"),
                      "desc": e.get("desc"), "amt": e.get("amt")} for e in e_picked],
    }


def _nova_new_id(arr):
    mx = 0
    for x in arr:
        try:
            mx = max(mx, int(x.get("id", 0) or 0))
        except Exception:
            pass
    return mx + 1


_ALLOWED_TASK = {"title", "status", "priority", "assignee", "assignees", "due", "notes", "type", "repeat", "time"}
_REPEATS = ("none", "daily", "weekdays", "weekly", "biweekly", "monthly")
_ASSIGNEE_IDS = ("nema", "arvin", "edgar")


def _norm_assignees(data):
    """Accept assignees[] (or a single assignee), keep only valid ids, dedup."""
    raw = data.get("assignees")
    if not isinstance(raw, list):
        raw = [data.get("assignee")] if data.get("assignee") else []
    out = []
    for a in raw:
        if a in _ASSIGNEE_IDS and a not in out:
            out.append(a)
    return out
_ALLOWED_DEAL = {"front", "back", "feeReferral", "feeProgram", "feeEnvy", "pay", "payBack", "lead", "source", "agentId",
                 "notes", "override", "type", "term", "dealer", "progPaid", "progPaidD", "refPaid", "refPaidD",
                 "envyColl", "envyCollD"}
# Expense categories in use (matches the imported ledger + dashboard ROAS matcher).
_EXPENSE_CATS = ("Ad Spend", "Software", "Office", "Developer", "Refunds", "Auto", "Other")
# Per-confirm action ceiling — big enough for a pasted expense list, small enough to
# bound one confirmed write. Overflow is REPORTED, never silently dropped.
_MAX_ACTIONS = 50


def _nova_apply_actions(store, actions, user=None):
    """Execute the agent's confirmed actions against the store IN PLACE. Pure &
    testable — every op is allow-listed and returns a per-action result."""
    uid = (user or {}).get("id") or "edgar"
    agents = store.get("agents", []) or []
    deals = store.setdefault("deals", [])
    tasks = store.setdefault("tasks", [])
    notes = store.setdefault("notes", [])
    expenses = store.setdefault("expenses", [])
    today = time.strftime("%Y-%m-%d")

    def jload(s):
        if isinstance(s, dict):
            return s
        try:
            return json.loads(s) if isinstance(s, str) and s.strip() else {}
        except Exception:
            return {}

    results = []
    acts = list(actions or [])
    if len(acts) > _MAX_ACTIONS:
        results.append({"op": "batch", "ok": False,
                        "error": f"only the first {_MAX_ACTIONS} of {len(acts)} actions were applied — send the rest in another message"})
        acts = acts[:_MAX_ACTIONS]
    for a in acts:
        op = a.get("op")
        data = jload(a.get("data"))
        rid = str(a.get("id") or "")
        try:
            if op == "create_task":
                _ids = _norm_assignees(data)
                t = {"id": _nova_new_id(tasks), "title": str(data.get("title") or "Untitled task"),
                     "status": data.get("status") if data.get("status") in ("backlog", "todo", "inprogress", "done") else "todo",
                     "priority": data.get("priority") if data.get("priority") in ("urgent", "high", "medium", "low", "none") else "none",
                     "assignees": _ids, "assignee": _ids[0] if _ids else "",
                     "due": data.get("due") or "", "notes": data.get("notes") or "",
                     "subtasks": [{"text": str(s), "done": False} for s in (data.get("subtasks") or [])],
                     "labels": [str(x) for x in (data.get("labels") or [])][:3], "dealId": None,
                     "type": "event" if data.get("type") == "event" else "task",
                     "repeat": data.get("repeat") if data.get("repeat") in _REPEATS else "none",
                     "time": data.get("time") or "", "createdBy": uid, "created": today}
                tasks.insert(0, t)
                results.append({"op": op, "ok": True, "id": t["id"], "label": t["title"]})
            elif op == "create_note":
                n = {"id": _nova_new_id(notes), "title": str(data.get("title") or "Untitled note"),
                     "blocks": [{"t": "p", "text": str(data.get("body") or "")}], "pinned": False,
                     "owner": uid, "sharedWith": [], "created": today, "updated": today}
                notes.insert(0, n)
                results.append({"op": op, "ok": True, "id": n["id"], "label": n["title"]})
            elif op == "create_deal":
                aid = data.get("agentId") if any(x.get("id") == data.get("agentId") for x in agents) else (agents[0].get("id") if agents else "")
                d = {"id": _nova_new_id(deals), "date": data.get("date") or today, "client": data.get("client") or "",
                     "year": data.get("year") or int(time.strftime("%Y")), "make": data.get("make") or "", "model": data.get("model") or "",
                     "vin": "", "dealer": data.get("dealer") or "", "type": data.get("type") if data.get("type") in ("Lease", "Buy") else "Lease",
                     "term": data.get("term") or "", "agentId": aid,
                     "lead": data.get("lead") if data.get("lead") in ("own", "nova", "referral") else "own",
                     "source": data.get("source") if data.get("source") in ("FB", "IG", "Google", "Yelp", "Referral", "Repeat", "Walk-in", "Website", "Other") else "",
                     "wOffered": False, "wSold": bool(data.get("wSold")), "front": _n_num(data.get("front")), "back": _n_num(data.get("back")),
                     "feeReferral": _n_num(data.get("feeReferral")),
                     "feeProgram": _n_num(data.get("feeProgram")),
                     # blank = auto 20% of back; a positive number = manual override
                     "feeEnvy": _n_num(data.get("feeEnvy")) if _n_num(data.get("feeEnvy")) > 0 else "",
                     "override": None,
                     "pay": data.get("pay") if data.get("pay") in ("Stripe", "Zelle", "Cash", "Check") else "Stripe",
                     "payBack": data.get("payBack") if data.get("payBack") in ("Zelle", "Check", "ACH", "Wire", "Cash") else "",
                     "fColl": False, "bColl": False, "aPaidF": False, "aPaidFd": "", "aPaidB": False, "aPaidBd": "",
                     # Shared-fee PAYOUTS (program/referral) can be marked paid at creation — the
                     # Garage note records them, e.g. "Program Fee: $750 (7/19 jason paid)". This is
                     # money OUT to the dealer/referrer, a different axis from the back-side money-IN
                     # rule (bColl/aPaidB stay hard-false above; Edgar confirms those himself).
                     "progPaid": bool(data.get("progPaid")),
                     "progPaidD": str(data.get("progPaidD") or "")[:10] if data.get("progPaid") else "",
                     "refPaid": bool(data.get("refPaid")),
                     "refPaidD": str(data.get("refPaidD") or "")[:10] if data.get("refPaid") else "",
                     "envyColl": False, "envyCollD": "",
                     "notes": data.get("notes") or ""}
                deals.insert(0, d)
                results.append({"op": op, "ok": True, "id": d["id"], "label": d["client"] or "deal"})
            elif op == "create_expense":
                amt = round(_n_num(data.get("amt")), 2)
                if amt <= 0:
                    results.append({"op": op, "ok": False, "error": "expense needs a positive amount"})
                    continue
                e = {"id": _nova_new_id(expenses),
                     "date": (str(data.get("date") or "") or today)[:10],
                     "cat": data.get("cat") if data.get("cat") in _EXPENSE_CATS else "Other",
                     "desc": str(data.get("desc") or "").strip()[:160],
                     "amt": amt}
                expenses.insert(0, e)
                results.append({"op": op, "ok": True, "id": e["id"],
                                "label": f"{e['cat']} · {e['desc'] or 'expense'} · ${amt:,.0f}"})
            elif op in ("update_expense", "delete_expense"):
                e = next((x for x in expenses if str(x.get("id")) == rid), None)
                if not e:
                    results.append({"op": op, "ok": False, "error": "expense not found"})
                    continue
                if op == "delete_expense":
                    expenses[:] = [x for x in expenses if x is not e]
                else:
                    for k, v in data.items():
                        if k == "amt":
                            e["amt"] = round(_n_num(v), 2)
                        elif k == "cat":
                            e["cat"] = v if v in _EXPENSE_CATS else e.get("cat", "Other")
                        elif k in ("date", "desc"):
                            e[k] = str(v or "")[:160]
                results.append({"op": op, "ok": True, "id": e.get("id"),
                                "label": f"{e.get('cat','')} · {e.get('desc') or 'expense'}"})
            elif op in ("update_task", "complete_task", "delete_task"):
                t = next((x for x in tasks if str(x.get("id")) == rid), None)
                if not t:
                    results.append({"op": op, "ok": False, "error": "task not found"})
                    continue
                if op == "delete_task":
                    tasks[:] = [x for x in tasks if x is not t]
                elif op == "complete_task":
                    t["status"] = "done"
                else:
                    for k, v in data.items():
                        if k in _ALLOWED_TASK and k not in ("assignee", "assignees"):
                            t[k] = v
                    if "assignees" in data or "assignee" in data:
                        _ids = _norm_assignees(data)
                        t["assignees"], t["assignee"] = _ids, (_ids[0] if _ids else "")
                results.append({"op": op, "ok": True, "id": t.get("id"), "label": t.get("title")})
            elif op in ("update_deal", "mark_deal_collected", "mark_agent_paid"):
                d = next((x for x in deals if str(x.get("id")) == rid), None)
                if not d:
                    results.append({"op": op, "ok": False, "error": "deal not found"})
                    continue
                if op == "update_deal":
                    for k, v in data.items():
                        if k in _ALLOWED_DEAL:
                            d[k] = v
                    # Keep fee paid-dates coherent (mirrors the UI toggle): a fee marked paid
                    # with no supplied date defaults to today; un-marking clears the date.
                    if "progPaid" in data:
                        d["progPaidD"] = (str(d.get("progPaidD") or "") or today) if d.get("progPaid") else ""
                    if "refPaid" in data:
                        d["refPaidD"] = (str(d.get("refPaidD") or "") or today) if d.get("refPaid") else ""
                elif op == "mark_deal_collected":
                    side = data.get("side", "both")
                    if side in ("front", "both"):
                        d["fColl"] = True
                    if side in ("back", "both"):
                        d["bColl"] = True
                else:  # mark_agent_paid
                    side = data.get("side", "both")
                    if side in ("front", "both"):
                        d["aPaidF"], d["aPaidFd"] = True, today
                    if side in ("back", "both"):
                        d["aPaidB"], d["aPaidBd"] = True, today
                results.append({"op": op, "ok": True, "id": d.get("id"), "label": d.get("client") or "deal"})
            else:
                results.append({"op": op, "ok": False, "error": "unknown op"})
        except Exception as e:
            results.append({"op": op, "ok": False, "error": str(e)[:120]})
    return results


_AGENT_ACTION = {
    "type": "object", "additionalProperties": False,
    "required": ["op", "id", "summary", "data"],
    "properties": {
        "op": {"type": "string", "enum": ["create_task", "create_deal", "create_note", "update_task",
                                          "complete_task", "update_deal", "mark_deal_collected",
                                          "mark_agent_paid", "delete_task",
                                          "create_expense", "update_expense", "delete_expense"]},
        "id": {"type": "string"},
        "summary": {"type": "string"},
        "data": {"type": "string"},
    },
}
_AGENT_SCHEMA = {
    "type": "object", "additionalProperties": False,
    "required": ["kind", "reply", "navigate_page", "navigate_hash", "actions"],
    "properties": {
        "kind": {"type": "string", "enum": ["answer", "navigate", "act"]},
        "reply": {"type": "string"},
        "navigate_page": {"type": "string", "enum": ["ledger", "tasks", "notes", ""]},
        "navigate_hash": {"type": "string"},
        "actions": {"type": "array", "items": _AGENT_ACTION},
    },
}


@app.route("/nova-admins/agent", methods=["POST"])
def nova_admins_agent():
    """The system-wide agent brain: reads a live snapshot, then answers, navigates,
    or PROPOSES writes (confirmed separately via /agent/apply)."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    body = request.json or {}
    text = (body.get("text") or "").strip()
    if not text:
        return jsonify({"error": "Ask me something."}), 400
    today = (body.get("today") or time.strftime("%Y-%m-%d")).strip()
    page = (body.get("page") or "").strip()
    me = _nova_current_user() or {"id": "edgar", "name": "Edgar"}
    with _NOVA_LOCK:
        store = _nova_load()
    snap = _nova_snapshot(store, today, text)
    sys = (
        f"You are the Nova Admins agent — a sharp, trusted back-office operator for Nova, a car brokerage. "
        f"You're talking to {me['name']} (id '{me['id']}') — when they say 'me'/'my'/'mine', that's '{me['id']}'. "
        "You are stickied on every admin page and see a LIVE snapshot of the business (headline metrics, agents "
        "and their split %, the most relevant deals, tasks, notes). Help from anywhere. Return STRICT JSON.\n"
        "Choose kind:\n"
        "- 'answer': they asked a question — answer concisely and SPECIFICALLY from the snapshot (real numbers, names, "
        "$ amounts). If the needed record isn't in the snapshot, say so briefly. No actions.\n"
        "- 'navigate': they want to go somewhere — set navigate_page (ledger|tasks|notes); navigate_hash only if sure "
        "(tasks views: '#view=board|list|table|cal|agenda'; ledger tabs open by default). reply = one short line.\n"
        "- 'act': they want to create or change records — PROPOSE actions (do not execute; the user confirms). "
        "reply = one short line describing what you'll do; give each action a human 'summary'.\n"
        "ACTIONS — 'data' MUST be a JSON string:\n"
        " create_task data={title, status, priority, assignees(array of nema/arvin/edgar), due(YYYY-MM-DD), repeat(none|daily|weekdays|weekly|biweekly|monthly), labels[], notes, subtasks[]}"
        " — for a calendar EVENT (a meeting/appointment) add type='event' and time='HH:MM' (24h); events show on the calendar.\n"
        " create_note data={title, body}\n"
        " create_deal data={client, year, make, model, dealer, type(Lease|Buy), agentId, lead(own|nova|referral, sets the split), source(channel for metrics: FB|IG|Google|Yelp|Referral|Repeat|Walk-in|Website|Other, '' if unknown), front, back, feeReferral, feeProgram, progPaid(bool), progPaidD(YYYY-MM-DD), refPaid(bool), refPaidD(YYYY-MM-DD), pay(FRONT method: Stripe|Zelle|Cash|Check), payBack(BACK method — dealer pays Nova: Check|ACH|Wire|Zelle|Cash, never Stripe, '' if no back), notes}"
        " — the Envy fee (20% of back) is AUTOMATIC when a back end exists; don't add it yourself. Stripe's 3% is auto on the FRONT only. "
        "If the note says a Program or Referral fee was ALREADY PAID — usually written right next to the fee, e.g. 'Program Fee: $750 (7/19 jason paid)' — set progPaid/refPaid=true and progPaidD/refPaidD to that date (resolve to YYYY-MM-DD). This is money OUT to the dealer/referrer, and is SEPARATE from collecting the deal's front/back money.\n"
        " update_task id=<taskId> data={status|priority|assignees(array)|due|title|notes}\n"
        " complete_task id=<taskId> data={}\n"
        " update_deal id=<dealId> data={front|back|feeReferral|feeProgram|feeEnvy|pay|payBack|lead|source|agentId|notes|override|progPaid|progPaidD|refPaid|refPaidD|envyColl}"
        " — Envy is a NOVA-ONLY COST: Envy receives the back, keeps 20% (Nova eats it), forwards 80% to Nova; it "
        "reduces Nova's take, never the agent's split. progPaid/refPaid = a shared fee was paid out; set progPaidD/refPaidD "
        "to the pay date (YYYY-MM-DD) when it's known, else it defaults to today.\n"
        " mark_deal_collected id=<dealId> data={side:'front'|'back'|'both'}\n"
        " mark_agent_paid id=<dealId> data={side:'front'|'back'|'both'}\n"
        " — BACK-SIDE rule for both mark ops: use side 'back'/'both' ONLY when Edgar explicitly says the back/dealer "
        "money came in (or the agent's back share was paid). An unqualified 'mark it collected/paid' means side:'front' "
        "— Edgar confirms back-end money himself.\n"
        " create_expense data={date(YYYY-MM-DD, today if unstated), cat(Ad Spend|Software|Office|Developer|Refunds|Auto|Other), desc, amt}"
        " — a business expense line (ad spend, tools, office…), one action per expense; amt = positive dollars. "
        "Pick the closest cat ('Other' if unclear); desc = short human label. Edgar may paste a whole LIST of "
        "expenses — emit one create_expense per line item (up to 50 per message), don't lump them into one.\n"
        " update_expense id=<expenseId> data={date|cat|desc|amt}\n"
        " delete_expense id=<expenseId> data={}\n"
        " delete_task id=<taskId> data={}\n"
        f"- Today is {today}; resolve relative dates to YYYY-MM-DD. assignees: nema/arvin/edgar ('me' = {me['id']}).\n"
        "- Match deals/tasks/agents by the ids in the snapshot. If no clearly-matching record exists for an update, "
        "return kind='answer' and say you couldn't find it — never guess an id.\n"
        "- Be decisive but safe: only propose what was clearly asked. Unused fields: navigate_page='', navigate_hash='', actions=[]."
    )
    user = "SNAPSHOT:\n" + json.dumps(snap) + "\n\nUSER REQUEST:\n" + text
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "system", "content": sys}, {"role": "user", "content": user}],
            response_format={"type": "json_schema", "json_schema": {"name": "agent", "strict": True, "schema": _AGENT_SCHEMA}},
        )
        out = json.loads(resp.choices[0].message.content)
        # enrich proposed new deals with the live split so the confirm card shows the money math
        if out.get("kind") == "act":
            amap = {a.get("id"): a for a in store.get("agents", [])}
            for a in out.get("actions") or []:
                if a.get("op") == "create_deal":
                    try:
                        d = json.loads(a.get("data") or "{}")
                        c = _nova_calc({"front": d.get("front"), "back": d.get("back"), "feeJason": d.get("feeJason"),
                                        "feeReferral": d.get("feeReferral"), "feeProgram": d.get("feeProgram"),
                                        "pay": d.get("pay"), "agentId": d.get("agentId"),
                                        "lead": d.get("lead"), "override": None}, amap)
                        a["summary"] = (a.get("summary") or "New deal") + \
                            f" — net ${round(c['net']):,} · agent ${round(c['agent']):,} · Nova ${round(c['nova']):,}"
                    except Exception:
                        pass
        _audit("agent_query", kind=out.get("kind"), page=page, user=me["id"], acts=len(out.get("actions") or []))
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": "Agent error — " + str(e)[:160]}), 500


@app.route("/nova-admins/agent/apply", methods=["POST"])
def nova_admins_agent_apply():
    """Execute the agent's CONFIRMED actions under the lock, with an audit line."""
    body = request.json or {}
    actions = body.get("actions") or []
    if not isinstance(actions, list) or not actions:
        return jsonify({"error": "No actions to apply."}), 400
    try:
        with _NOVA_LOCK:
            store = _nova_load()
            results = _nova_apply_actions(store, actions, _nova_current_user())
            _nova_write(store)
        _audit("agent_apply", ops=",".join(r.get("op", "") for r in results),
               ok=sum(1 for r in results if r.get("ok")), fail=sum(1 for r in results if not r.get("ok")))
        return jsonify({"ok": True, "results": results})
    except Exception as e:
        return jsonify({"error": str(e)[:160]}), 500


@app.route("/nova-admins/agent.js")
def nova_admins_agent_js():
    """The shared, stickied agent widget, injected into every Admin page."""
    return app.response_class(open(os.path.join(BASE_DIR, "nova_agent.js"), encoding="utf-8").read(),
                              mimetype="application/javascript")


def _note_img_dir():
    return os.path.join(GENERATED_DIR, "nova_note_images")


@app.route("/nova-admins/common.css")
def nova_admins_common_css():
    """Shared Nova Admins stylesheet (palette + base components), cached across pages."""
    resp = app.response_class(open(os.path.join(BASE_DIR, "nova_common.css"), encoding="utf-8").read(),
                              mimetype="text/css")
    return resp


@app.route("/nova-admins/common.js")
def nova_admins_common_js():
    """Shared Nova Admins helpers (esc, storage, row saves), cached across pages."""
    return app.response_class(open(os.path.join(BASE_DIR, "nova_common.js"), encoding="utf-8").read(),
                              mimetype="application/javascript")


@app.route("/nova-admins/img", methods=["POST"])
def nova_admins_img_upload():
    """Store a note image ONCE (so autosaves don't re-ship base64). Returns its URL."""
    body = request.get_json(silent=True) or {}
    m = re.match(r"^data:image/(png|jpeg|jpg|webp|gif);base64,(.+)$", body.get("data") or "", re.DOTALL)
    if not m:
        return jsonify({"error": "Expected a base64 image data URL."}), 400
    ext = {"jpeg": "jpg"}.get(m.group(1), m.group(1))
    try:
        raw = base64.b64decode(m.group(2))
    except Exception:
        return jsonify({"error": "Bad image data."}), 400
    if len(raw) > 8_000_000:
        return jsonify({"error": "Image too large (8MB max)."}), 413
    os.makedirs(_note_img_dir(), exist_ok=True)
    name = secrets.token_hex(12) + "." + ext
    with open(os.path.join(_note_img_dir(), name), "wb") as f:
        f.write(raw)
    _audit("note_img_upload", bytes=len(raw))
    return jsonify({"url": "/nova-admins/img/" + name})


@app.route("/nova-admins/img/<name>")
def nova_admins_img_get(name):
    """Serve a stored note image (gated by the admin session; no path traversal)."""
    if not re.match(r"^[a-f0-9]{24}\.(png|jpg|jpeg|webp|gif)$", name):
        return "", 404
    path = os.path.join(_note_img_dir(), name)
    return send_file(path) if os.path.exists(path) else ("", 404)


@app.route("/toolbox")
def toolbox():
    return send_file(os.path.join(BASE_DIR, "toolbox.html"))


@app.route("/toolbox/pdf")
def toolbox_pdf_page():
    return send_file(os.path.join(BASE_DIR, "toolbox_pdf.html"))


@app.route("/toolbox/rx")
def toolbox_rx_page():
    return send_file(os.path.join(BASE_DIR, "toolbox_rx.html"))


# ---------------------------------------------------------------------------
# Aletheya Toolbox · PDF Filler
# Reads a fillable PDF (AcroForm) template, fills its named fields, and exports.
# Three ways in: type values, AI-parse pasted text, or batch from a spreadsheet.
# ---------------------------------------------------------------------------

def _pdf_field_info(reader):
    """Return a clean list of the template's form fields for the UI."""
    out = []
    fields = reader.get_fields() or {}
    for name, f in fields.items():
        ftype = f.get("/FT")
        kind = {"/Tx": "text", "/Btn": "checkbox", "/Ch": "choice", "/Sig": "signature"}.get(ftype, "text")
        info = {"name": name, "type": kind}
        # Checkbox/radio: surface the "on" state(s) so we know what value turns it on.
        states = f.get("/_States_")
        if states:
            info["states"] = [str(s) for s in states]
        # Dropdown / list options.
        opts = f.get("/Opt") or f.get("/_States_")
        if kind == "choice" and opts:
            info["options"] = [str(o[1]) if isinstance(o, (list, tuple)) else str(o) for o in opts]
        cur = f.get("/V")
        if cur is not None:
            info["value"] = str(cur)
        out.append(info)
    return out


def _fill_pdf(template_bytes, values):
    """Fill every page's form fields with `values` (name -> value) and return bytes."""
    from pypdf import PdfReader, PdfWriter
    reader = PdfReader(io.BytesIO(template_bytes))
    writer = PdfWriter()
    writer.append(reader)
    # Coerce booleans to the checkbox "on" state where we can detect it.
    fields = reader.get_fields() or {}
    clean = {}
    for k, v in values.items():
        if k not in fields:
            continue
        f = fields[k]
        if f.get("/FT") == "/Btn":
            states = [str(s) for s in (f.get("/_States_") or [])]
            on = next((s for s in states if s not in ("/Off", "Off")), "/Yes")
            if isinstance(v, bool):
                clean[k] = on if v else "/Off"
            elif str(v).strip().lower() in ("1", "true", "yes", "x", "on", "checked", "✓"):
                clean[k] = on
            elif str(v).strip().lower() in ("0", "false", "no", "off", "", "unchecked"):
                clean[k] = "/Off"
            else:
                clean[k] = str(v)
        else:
            clean[k] = "" if v is None else str(v)
    for page in writer.pages:
        try:
            writer.update_page_form_field_values(page, clean, auto_regenerate=False)
        except Exception:
            pass
    # Make viewers render the filled values without re-saving.
    try:
        writer.set_need_appearances_writer(True)
    except Exception:
        pass
    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()


@app.route("/toolbox/pdf-fields", methods=["POST"])
def toolbox_pdf_fields():
    """Inspect an uploaded PDF template and return its fillable fields."""
    f = request.files.get("template")
    if not f:
        return jsonify({"error": "No PDF uploaded."}), 400
    try:
        from pypdf import PdfReader
        data = f.read()
        reader = PdfReader(io.BytesIO(data))
        fields = _pdf_field_info(reader)
    except Exception as e:
        return jsonify({"error": f"Couldn't read that PDF: {e}"}), 400
    if not fields:
        return jsonify({"fields": [], "message": "This PDF has no fillable form fields. Use a PDF with form fields (AcroForm) made in Acrobat or similar."})
    return jsonify({"fields": fields, "pages": len(reader.pages)})


@app.route("/toolbox/pdf-fill", methods=["POST"])
def toolbox_pdf_fill():
    """Fill a single PDF from a JSON map of field values and return the PDF."""
    f = request.files.get("template")
    if not f:
        return jsonify({"error": "No PDF uploaded."}), 400
    try:
        values = json.loads(request.form.get("values", "{}"))
    except Exception:
        return jsonify({"error": "Bad values payload."}), 400
    try:
        out = _fill_pdf(f.read(), values or {})
    except Exception as e:
        return jsonify({"error": f"Couldn't fill the PDF: {e}"}), 500
    name = (request.form.get("filename") or "filled").strip() or "filled"
    if not name.lower().endswith(".pdf"):
        name += ".pdf"
    return send_file(io.BytesIO(out), mimetype="application/pdf",
                     as_attachment=True, download_name=name)


@app.route("/toolbox/pdf-parse", methods=["POST"])
def toolbox_pdf_parse():
    """AI-map free-form pasted text onto the template's field names."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    data = request.json or {}
    fields = data.get("fields") or []
    text = (data.get("text") or "").strip()
    if not fields:
        return jsonify({"error": "No template fields provided."}), 400
    if not text:
        return jsonify({"error": "Paste some text to parse."}), 400
    # Describe the fields so the model knows the exact keys to emit.
    lines = []
    for f in fields:
        d = f"- {f['name']} ({f.get('type','text')}"
        if f.get("options"):
            d += "; options: " + ", ".join(f["options"])
        elif f.get("states"):
            d += "; on-state: " + (f["states"][0] if f["states"] else "/Yes")
        d += ")"
        lines.append(d)
    field_list = "\n".join(lines)
    prompt = (
        "You map free-form information onto a fixed set of PDF form fields.\n"
        "Return ONLY a JSON object whose keys are EXACT field names from the list below.\n"
        "Only include a key when you can confidently determine its value from the text; "
        "omit fields you cannot fill. For checkboxes, use true/false. For dropdowns/choices, "
        "use one of the listed options verbatim. Do not invent data.\n\n"
        f"FIELDS:\n{field_list}\n\nTEXT:\n{text}"
    )
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        out = json.loads(resp.choices[0].message.content)
    except Exception as e:
        return jsonify({"error": f"AI parse failed: {e}"}), 500
    names = {f["name"] for f in fields}
    values = {k: v for k, v in out.items() if k in names}
    return jsonify({"values": values})


def _read_spreadsheet(file_storage):
    """Parse an uploaded CSV or XLSX into (headers, list-of-row-dicts)."""
    name = (file_storage.filename or "").lower()
    raw = file_storage.read()
    rows = []
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        grid = [[("" if c is None else c) for c in r] for r in ws.iter_rows(values_only=True)]
    else:
        text = raw.decode("utf-8-sig", "ignore")
        grid = list(csv.reader(io.StringIO(text)))
    grid = [r for r in grid if any(str(c).strip() for c in r)]
    if not grid:
        return [], []
    headers = [str(h).strip() for h in grid[0]]
    for r in grid[1:]:
        row = {headers[i]: (str(r[i]).strip() if i < len(r) and r[i] is not None else "")
               for i in range(len(headers))}
        rows.append(row)
    return headers, rows


@app.route("/toolbox/pdf-batch", methods=["POST"])
def toolbox_pdf_batch():
    """Fill the template once per spreadsheet row; return a zip of PDFs.

    Spreadsheet column headers are matched to PDF field names (exact, then
    case-insensitive). An optional 'filename' column names each output PDF.
    """
    tpl = request.files.get("template")
    sheet = request.files.get("sheet")
    if not tpl or not sheet:
        return jsonify({"error": "Upload both a PDF template and a spreadsheet."}), 400
    template_bytes = tpl.read()
    try:
        from pypdf import PdfReader
        field_names = list((PdfReader(io.BytesIO(template_bytes)).get_fields() or {}).keys())
    except Exception as e:
        return jsonify({"error": f"Couldn't read the PDF template: {e}"}), 400
    if not field_names:
        return jsonify({"error": "The PDF template has no fillable form fields."}), 400
    try:
        headers, rows = _read_spreadsheet(sheet)
    except Exception as e:
        return jsonify({"error": f"Couldn't read the spreadsheet: {e}"}), 400
    if not rows:
        return jsonify({"error": "The spreadsheet has no data rows."}), 400
    # Map headers -> field names (exact, else case-insensitive).
    lower = {fn.lower(): fn for fn in field_names}
    colmap = {}
    for h in headers:
        if h in field_names:
            colmap[h] = h
        elif h.lower() in lower:
            colmap[h] = lower[h.lower()]
    if not colmap:
        return jsonify({"error": "No spreadsheet columns matched the PDF field names. "
                                 "Make the column headers match the field names exactly. "
                                 f"PDF fields: {', '.join(field_names)}"}), 400
    base = (request.form.get("filename") or "filled").strip() or "filled"
    zbuf = io.BytesIO()
    used = set()
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as z:
        for i, row in enumerate(rows, 1):
            values = {colmap[h]: row[h] for h in colmap if row.get(h, "") != ""}
            try:
                pdf = _fill_pdf(template_bytes, values)
            except Exception:
                continue
            fn = (row.get("filename") or row.get("Filename") or f"{base}-{i}").strip()
            fn = re.sub(r"[^\w.\- ]+", "", fn) or f"{base}-{i}"
            if not fn.lower().endswith(".pdf"):
                fn += ".pdf"
            if fn in used:
                fn = f"{fn[:-4]}-{i}.pdf"
            used.add(fn)
            z.writestr(fn, pdf)
    zbuf.seek(0)
    return send_file(zbuf, mimetype="application/zip",
                     as_attachment=True, download_name=f"{base}-batch.zip")


# ---------------------------------------------------------------------------
# Aletheya Toolbox · RX Update (Medical / Psychiatric Update) generator
# ---------------------------------------------------------------------------

@app.route("/toolbox/doctors")
def toolbox_doctors():
    """The provider directory — used to populate the doctor picker."""
    out = []
    for i, d in enumerate(rx_form.DOCTORS):
        out.append({"index": i, "label": rx_form.doctor_label(d), **d})
    return jsonify({"doctors": out})


@app.route("/toolbox/rx-parse", methods=["POST"])
def toolbox_rx_parse():
    """Read a chart screenshot and/or pasted text into structured RX-Update data."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    data = request.json or {}
    text = (data.get("text") or "").strip()
    images = data.get("images") or []          # list of data URLs
    instructions = (data.get("instructions") or "").strip()
    if not text and not images:
        return jsonify({"error": "Add a screenshot or some text to parse."}), 400

    sys = (
        "You extract patient data from an Adult Day Health Care chart (e.g. TurboADHC "
        "Plan of Care) for a Medical/Psychiatric Update fax. Return ONLY a JSON object "
        "with these keys:\n"
        '  patient_name (string, "First Last"),\n'
        "  dob (string, MM/DD/YYYY),\n"
        "  diagnoses (array of {name, icd}) — use the ICD-10 code exactly as shown,\n"
        "  medications (array of strings — full sig if shown, e.g. 'Aspirin 81 mg QD PO'),\n"
        "  significant_events (string, may be empty).\n"
        "Read carefully from the image(s) and text. Do not invent data — omit what you can't read. "
        "If the user gives instructions (e.g. remove or add a medication, fix a name), follow them."
    )
    user_content = []
    parts = []
    if text:
        parts.append("SOURCE TEXT:\n" + text)
    if instructions:
        parts.append("INSTRUCTIONS:\n" + instructions)
    if not parts:
        parts.append("Extract everything you can from the attached image(s).")
    user_content.append({"type": "text", "text": "\n\n".join(parts)})
    for url in images[:6]:
        if isinstance(url, str) and url.startswith("data:image"):
            user_content.append({"type": "image_url", "image_url": {"url": url}})

    try:
        client = _oai()
        resp = client.chat.completions.create(
            model="gpt-4.1",
            messages=[{"role": "system", "content": sys},
                      {"role": "user", "content": user_content}],
            response_format={"type": "json_object"},
        )
        out = json.loads(resp.choices[0].message.content)
    except Exception as e:
        return jsonify({"error": f"AI parse failed: {e}"}), 500

    # Normalize shape.
    result = {
        "patient_name": str(out.get("patient_name") or "").strip(),
        "dob": str(out.get("dob") or "").strip(),
        "diagnoses": [],
        "medications": [],
        "significant_events": str(out.get("significant_events") or "").strip(),
    }
    for dx in (out.get("diagnoses") or []):
        if isinstance(dx, dict):
            name = str(dx.get("name") or dx.get("diagnosis") or "").strip()
            icd = str(dx.get("icd") or dx.get("icd_code") or dx.get("code") or "").strip()
        else:
            name, icd = str(dx).strip(), ""
        if name or icd:
            result["diagnoses"].append({"name": name, "icd": icd})
    for m in (out.get("medications") or []):
        m = (m if isinstance(m, str) else str(m.get("name", "")) if isinstance(m, dict) else str(m)).strip()
        if m:
            result["medications"].append(m)
    return jsonify(result)


@app.route("/toolbox/rx-export", methods=["POST"])
def toolbox_rx_export():
    """Render the RX Update to PDF or Word from the structured data."""
    data = request.json or {}
    fmt = (data.get("format") or "pdf").lower()
    payload = data.get("data") or {}
    try:
        if fmt == "docx":
            blob = rx_form.render_rx_docx(payload)
            mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            ext = ".docx"
        else:
            blob = rx_form.render_rx_pdf(payload)
            mime = "application/pdf"
            ext = ".pdf"
    except Exception as e:
        return jsonify({"error": f"Couldn't build the document: {e}"}), 500
    base = (data.get("filename") or "").strip()
    if not base:
        pn = (payload.get("patient_name") or "RX-Update").strip().replace(" ", "-")
        base = f"RX-Update-{pn}" if pn else "RX-Update"
    base = re.sub(r"[^\w.\-]+", "", base) or "RX-Update"
    if base.lower().endswith(ext):
        base = base[:-len(ext)]
    return send_file(io.BytesIO(blob), mimetype=mime, as_attachment=True, download_name=base + ext)


@app.route("/status")
def status():
    return jsonify({"has_key": bool(API_KEY and API_KEY != "sk-your-key-here")})


@app.route("/generate", methods=["POST"])
def generate():
    data = request.json
    vehicle = data.get("vehicle", "").strip()
    bodystyle = data.get("bodystyle", "").strip()
    color = data.get("color", "").strip()
    prompt_template = data.get("prompt_template", "").strip()
    do_research = data.get("research", True)
    force = bool(data.get("force", False))

    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500

    if not vehicle:
        return jsonify({"error": "Missing vehicle name."}), 400

    # Deterministic identity: same vehicle + body style + color -> same file.
    filename = f"{car_slug(vehicle, bodystyle, color)}.png"
    filepath = os.path.join(GENERATED_DIR, filename)

    # FAIL-SAFE: if this exact car was already generated, reuse it (no API call).
    if not force and os.path.exists(filepath):
        return jsonify({
            "success": True,
            "filename": filename,
            "vehicle": vehicle,
            "already_existed": True,
        })

    try:
        client = _oai()

        # STEP 1 — look up the real, current design online (so a brand-new model year
        # is rendered correctly instead of an outdated training-data version).
        reference = research_vehicle(client, vehicle, bodystyle) if do_research else ""

        # STEP 2 — build the image prompt (optional custom override still supported).
        if prompt_template:
            prompt = (prompt_template
                      .replace("{vehicle}", vehicle)
                      .replace("{bodystyle}", bodystyle or "as per the actual production model")
                      .replace("{color}", color or "the standard factory color for this model"))
            if reference:
                prompt += ("\n\nAUTHORITATIVE REAL-WORLD REFERENCE (match this exact model year):\n"
                           + reference)
        else:
            prompt = build_image_prompt(vehicle, bodystyle, color, reference)

        result = client.images.generate(
            model="gpt-image-1",
            prompt=prompt,
            n=1,
            size="1536x1024",
            quality="high",
            background="transparent",
            output_format="png",
        )

        image_b64 = result.data[0].b64_json
        image_bytes = base64.b64decode(image_b64)

        with open(filepath, "wb") as f:
            f.write(image_bytes)

        return jsonify({
            "success": True,
            "filename": filename,
            "vehicle": vehicle,
            "already_existed": False,
            "referenced": bool(reference),
            "reference": reference,
        })

    except Exception as e:
        info = classify_error(e)
        print(f"[generate] error: {info['reason']} | {info['detail']}")
        return jsonify({
            "error": info["detail"] or info["reason"],
            "reason": info["reason"],
            "fatal": info["fatal"],
            "code": info["code"],
        }), info["status"]


@app.route("/images/<path:filename>")
def serve_image(filename):
    return send_from_directory(GENERATED_DIR, filename)


@app.route("/images")
def list_images():
    files = sorted(os.listdir(GENERATED_DIR), reverse=True)
    # exclude generated backgrounds (bg_*) from the car gallery
    files = [f for f in files if f.endswith(".png") and not f.startswith("bg_")]
    return jsonify(files)


BG_PROMPTS = {
    "studio": (
        "A premium empty studio backdrop for a vertical 9:16 poster. Deep navy-to-black "
        "gradient walls, a soft cool-blue glow rising from the bottom-center like stage "
        "lighting, a subtle glossy reflective floor along the bottom, faint sharp geometric "
        "panel seams on the side walls, dark, cinematic, high-end, minimal. Keep the upper-"
        "middle area darker and clear for overlaying content. Absolutely no text, no logos, "
        "no people, no cars, no products."
    ),
    "nova": (
        "A premium vertical 9:16 poster backdrop in a bold geometric brand style. Dark navy "
        "background with large angular triangular facets and paneling, crisp electric-blue "
        "edge glow and accents, a cool-blue light beam rising from the bottom-center, a subtle "
        "reflective floor, cinematic and modern. Keep the center clear and darker for overlaying "
        "content. Absolutely no text, no logos, no people, no cars, no products."
    ),
}


def bg_prompt(style):
    base = BG_PROMPTS.get(style, BG_PROMPTS["studio"])
    variants = ("", " cooler tones", " a tighter light beam", " a softer haze",
                " a wider glow", " deeper shadows", " faint volumetric fog", " a brighter floor reflection")
    return base + " Variation:" + variants[os.urandom(1)[0] % len(variants)] + "."


@app.route("/caption", methods=["POST"])
def caption():
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    data = request.json or {}
    vehicle = data.get("vehicle", "").strip()
    badge = data.get("badge", "").strip()
    if not vehicle:
        return jsonify({"error": "Add a make/model first."}), 400
    try:
        client = _oai()
        prompt = (
            "Write a short, punchy Instagram caption for a luxury car dealership called Nova Auto.\n"
            f"Context: {badge or 'SOLD'} — {vehicle}.\n"
            "Tone: upscale, celebratory, confident, concise. 1-2 sentences, then a new line with "
            "4-7 relevant hashtags. Include 1-2 tasteful emojis. No quotation marks around the caption. "
            "Vary the wording each time and keep it authentic to a high-end dealership."
        )
        resp = client.responses.create(model="gpt-4.1-mini", input=prompt)
        text = (getattr(resp, "output_text", "") or "").strip()
        return jsonify({"success": True, "caption": text})
    except Exception as e:
        info = classify_error(e)
        print(f"[caption] error: {info['reason']} | {info['detail']}")
        return jsonify({"error": info["detail"] or info["reason"], "reason": info["reason"]}), info["status"]


@app.route("/review-parse", methods=["POST"])
def review_parse():
    """Smart-fill: read a raw review pasted from Google/Yelp and pull out the reviewer
    name, star rating, source, and a clean version of the text for an Instagram graphic."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    text = ((request.json or {}).get("text") or "").strip()
    if not text:
        return jsonify({"error": "Paste a review first."}), 400
    schema = {
        "type": "object", "additionalProperties": False,
        "required": ["name", "rating", "source", "review", "vehicle"],
        "properties": {
            "name": {"type": "string"},
            "rating": {"type": "integer"},
            "source": {"type": "string"},
            "review": {"type": "string"},
            "vehicle": {"type": "string"},
        },
    }
    system = (
        "You read a customer review copied from Google, Yelp, DealerRater, or Cars.com and structure "
        "it for a social-media testimonial graphic. Output: name (the reviewer's name or first name + "
        "last initial like 'John D.'; '' if unknown), rating (1-5 integer; if not stated, 5), source "
        "(Google / Yelp / DealerRater / Cars.com / '' if unknown), review (clean the text up for a "
        "post: fix obvious typos/caps, trim filler, keep it authentic and in the customer's voice, no "
        "longer than ~320 characters; do NOT invent content), and vehicle (the car they bought if "
        "mentioned, e.g. '2024 BMW X5'; else '')."
    )
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "system", "content": system}, {"role": "user", "content": text}],
            response_format={"type": "json_schema", "json_schema": {"name": "review", "strict": True, "schema": schema}},
        )
        out = json.loads(resp.choices[0].message.content)
        r = out.get("rating") or 5
        out["rating"] = max(1, min(5, int(r) if str(r).isdigit() else 5))
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": f"Couldn't read that review: {e}"}), 502


@app.route("/reviews-parse", methods=["POST"])
def reviews_parse():
    """Bulk smart-fill: split a paste of SEVERAL reviews into a structured list, one
    object per review, for batch testimonial graphics."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    text = ((request.json or {}).get("text") or "").strip()
    if not text:
        return jsonify({"error": "Paste some reviews first."}), 400
    review_obj = {
        "type": "object", "additionalProperties": False,
        "required": ["name", "rating", "source", "review", "vehicle"],
        "properties": {
            "name": {"type": "string"}, "rating": {"type": "integer"},
            "source": {"type": "string"}, "review": {"type": "string"}, "vehicle": {"type": "string"},
        },
    }
    schema = {"type": "object", "additionalProperties": False, "required": ["reviews"],
              "properties": {"reviews": {"type": "array", "items": review_obj}}}
    system = (
        "You read a block of text that contains ONE OR MORE customer reviews (copied from Google, Yelp, "
        "DealerRater, or Cars.com — possibly several stacked together) and split them into a list. For "
        "EACH distinct review output: name (reviewer's name or first name + last initial like 'John D.'; "
        "'' if unknown), rating (1-5 integer; 5 if not stated), source (Google / Yelp / DealerRater / "
        "Cars.com / '' if unknown), review (cleaned up for a post: fix typos/caps, trim filler, keep it "
        "authentic and in the customer's voice, ≤320 chars; do NOT invent content), and vehicle (the car "
        "if mentioned, e.g. '2024 BMW X5'; else ''). Do not merge separate reviews; do not fabricate reviews."
    )
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "system", "content": system}, {"role": "user", "content": text}],
            response_format={"type": "json_schema", "json_schema": {"name": "reviews", "strict": True, "schema": schema}},
        )
        out = json.loads(resp.choices[0].message.content)
        rows = out.get("reviews") or []
        for r in rows:
            v = r.get("rating") or 5
            r["rating"] = max(1, min(5, int(v) if str(v).isdigit() else 5))
        return jsonify({"reviews": rows})
    except Exception as e:
        return jsonify({"error": f"Couldn't read those reviews: {e}"}), 502


@app.route("/review-caption", methods=["POST"])
def review_caption():
    """Write an Instagram caption + hashtags for a customer-review repost."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    d = request.json or {}
    review = (d.get("review") or "").strip()
    name = (d.get("name") or "").strip()
    rating = d.get("rating") or 5
    vehicle = (d.get("vehicle") or "").strip()
    source = (d.get("source") or "").strip()
    if not review:
        return jsonify({"error": "Add the review text first."}), 400
    prompt = (
        "Write a short, warm Instagram caption for a luxury car dealership, NovAuto, that is "
        "RESHARING a happy customer's review/testimonial.\n"
        f"Reviewer: {name or 'a happy customer'}. Rating: {rating}/5 stars"
        + (f" on their {vehicle}" if vehicle else "")
        + (f". Left on {source}" if source else "") + ".\n"
        f'The review: "{review}"\n'
        "Tone: grateful, celebratory, upscale, authentic. Thank the customer and warmly invite others in. "
        "1-3 short sentences with 1-2 tasteful emojis. Then a blank line, then a line of 8-12 relevant "
        "hashtags (mix brand, local car-buying/dealership, and review/testimonial tags). Do NOT wrap the "
        "caption in quotation marks. Vary the wording each time."
    )
    try:
        client = _oai()
        resp = client.responses.create(model="gpt-4.1-mini", input=prompt)
        text = (getattr(resp, "output_text", "") or "").strip()
        return jsonify({"caption": text})
    except Exception as e:
        info = classify_error(e)
        return jsonify({"error": info["detail"] or info["reason"], "reason": info["reason"]}), info["status"]


def _carsxe_dataurl(url, tries=3, timeout=20):
    """Fetch an image and return it as a data URL. The photo CDN can be slow and some
    hosts reject bare requests, so send browser-like headers and retry before giving up."""
    last = None
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
        "Accept": "image/avif,image/webp,image/png,image/jpeg,image/*,*/*;q=0.8",
        "Referer": "https://api.carsxe.com/",
    }
    for _ in range(tries):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=timeout) as r:
                ctype = r.headers.get("Content-Type", "image/png")
                b = r.read()
            return f"data:{ctype};base64," + base64.b64encode(b).decode()
        except Exception as e:
            last = e
    raise last


def _rgb_to_hex(rgb):
    """'226,5,0' -> '#e20500'. Returns '' if it can't parse."""
    try:
        parts = [max(0, min(255, int(float(x)))) for x in str(rgb).split(",")[:3]]
        if len(parts) == 3:
            return "#%02x%02x%02x" % tuple(parts)
    except Exception:
        pass
    return ""


def _carsxe_ymm(make, model, year, trim):
    """One /v1/ymm lookup -> (colors, trims). Empty lists on any failure."""
    params = {"key": CARSXE_API_KEY, "make": make, "model": model, "allTrimOptions": "1", "format": "json"}
    if year:
        params["year"] = year
    if trim:
        params["trim"] = trim
    url = "https://api.carsxe.com/v1/ymm?" + urllib.parse.urlencode(params)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as r:
            body = json.loads(r.read().decode("utf-8", "ignore"))
    except Exception:
        return [], []
    best = body.get("bestMatch") or body.get("data") or {}
    cands = []
    col = best.get("color") or best.get("colors") or {}
    if isinstance(col, dict):
        cands.append(col.get("exterior"))
    elif isinstance(col, list):
        cands.append(col)
    cands += [best.get("exterior_colors"), body.get("colors"), body.get("exterior_colors")]
    ext = next((c for c in cands if isinstance(c, list) and c), [])
    colors, seen = [], set()
    for c in ext:
        if not isinstance(c, dict):
            continue
        name = (c.get("name") or c.get("color_name") or c.get("description") or "").strip()
        if name and name.lower() not in seen:
            seen.add(name.lower())
            colors.append({"name": name, "hex": _rgb_to_hex(c.get("rgb") or c.get("rgb_value") or "")})
    trims = []
    for t in (body.get("trimOptions") or best.get("trimOptions") or []):
        nm = t if isinstance(t, str) else (t.get("trim") or t.get("name") or "")
        nm = (nm or "").strip()
        if nm and nm not in trims:
            trims.append(nm)
    return colors, trims


@app.route("/carsxe-meta", methods=["POST"])
def carsxe_meta():
    """Manufacturer exterior colors + trims for a vehicle. Falls back (drop trim,
    then recent prior years) when the exact year has no colour data, so the colour
    picker reliably shows the factory colours."""
    if not CARSXE_API_KEY:
        return jsonify({"error": "CARSXE_API_KEY is not set on the server."}), 500
    data = request.json or {}
    make = data.get("make", "").strip()
    model = _carsxe_base_model(data.get("model", "").strip())   # "Sorento Hybrid" -> "Sorento" for the catalog
    year = str(data.get("year", "")).strip()
    trim = data.get("trim", "").strip()
    if not (make and model):
        return jsonify({"colors": [], "trims": []})
    mkey = "|".join([year, make.lower(), model.lower(), trim.lower()])
    cached = _CARSXE_META_CACHE.get(mkey)
    if cached is not None:
        return jsonify({**cached, "cached": True})

    # exact -> drop trim -> recent prior years (color data lags for brand-new years)
    candidates = [(year, trim)]
    if trim:
        candidates.append((year, ""))
    try:
        yr = int(year)
        for back in (1, 2, 3):
            candidates.append((str(yr - back), ""))
    except Exception:
        pass

    colors, trims, tried = [], [], set()
    for (y, tr) in candidates:
        if (y, tr) in tried:
            continue
        tried.add((y, tr))
        c, t = _carsxe_ymm(make, model, y, tr)
        if t and not trims:
            trims = t
        if c:
            colors = c
            break

    payload = {"colors": colors, "trims": trims}
    if colors or trims:
        _cache_put(_CARSXE_META_CACHE, mkey, payload)
    return jsonify(payload)


def _carsxe_canon(make, model, year, trim):
    """Resolve messy make/model/trim against CarsXE /v1/ymm 'bestMatch' so the Deal
    Hub can snap a dealer's free-text car to the manufacturer's official spelling.
    Returns {found, make, model, trim, name, keys} — '' for any field it can't read.
    'keys' lists the bestMatch field names so we can adjust the field-picking if
    CarsXE names things differently than expected."""
    params = {"key": CARSXE_API_KEY, "make": make, "model": model, "format": "json"}
    if year:
        params["year"] = year
    if trim:
        params["trim"] = trim
    url = "https://api.carsxe.com/v1/ymm?" + urllib.parse.urlencode(params)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=20) as r:
            body = json.loads(r.read().decode("utf-8", "ignore"))
    except Exception as e:
        return {"found": False, "error": str(e)[:140]}
    best = body.get("bestMatch") or body.get("data") or {}
    if not isinstance(best, dict):
        best = {}
    attrs = best.get("attributes") if isinstance(best.get("attributes"), dict) else {}

    def pick(*names):
        for src in (best, attrs):
            for n in names:
                v = src.get(n)
                if isinstance(v, str) and v.strip():
                    return v.strip()
                if isinstance(v, (int, float)):
                    return str(v)
        return ""

    keys = list(best.keys())[:30] + (["attributes." + k for k in list(attrs.keys())[:25]] if attrs else [])
    try:
        raw = json.dumps(best, default=str)[:2000]
    except Exception:
        raw = str(best)[:2000]
    return {
        "found": bool(best),
        "make": pick("make", "make_name", "manufacturer", "brand"),
        "model": pick("model", "model_name"),
        "trim": pick("trim", "trim_name", "trim_level", "trim_description", "style", "style_name"),
        "name": pick("name", "full_name", "vehicle", "description", "vehicle_name"),
        "keys": keys,
        "body_keys": list(body.keys())[:20],   # where the canonical fields actually live, for diagnosis
        "raw": raw,                            # the full bestMatch object so we can see real field names/values
    }


@app.route("/carsxe-canon", methods=["POST"])
def carsxe_canon():
    """Batch name-cleanup: snap each {year,make,model,trim} to CarsXE's official
    spelling. Bounded + cached so one request can't run forever or re-hit the API."""
    if not CARSXE_API_KEY:
        return jsonify({"error": "CARSXE_API_KEY is not set on the server."}), 500
    data = request.json or {}
    cars = data.get("cars") or []
    if not isinstance(cars, list) or not cars:
        return jsonify({"results": []})
    capped = len(cars) > 40
    out = []
    for c in cars[:40]:
        make = (c.get("make") or "").strip()
        model = (c.get("model") or "").strip()
        year = str(c.get("year") or "").strip()
        trim = (c.get("trim") or "").strip()
        if not (make and model):
            out.append({"in": c, "found": False})
            continue
        mkey = "|".join([year, make.lower(), model.lower(), trim.lower()])
        res = _CARSXE_CANON_CACHE.get(mkey)
        if res is None:
            res = _carsxe_canon(make, model, year, trim)
            if res.get("found"):
                _cache_put(_CARSXE_CANON_CACHE, mkey, res)
        out.append({"in": c, **res})
    return jsonify({"results": out, "capped": capped})


@app.route("/carsxe-image", methods=["POST"])
def carsxe_image():
    """TRIAL: pull a vehicle photo from the catalog's /images endpoint (make/model based,
    transparent by default) so we can compare its image quality against Vehicle
    Databases. Returns the full set of results so the user can pick the best one."""
    if not CARSXE_API_KEY:
        return jsonify({"error": "CARSXE_API_KEY is not set on the server. Add it in Railway → Variables."}), 500
    data = request.json or {}
    make = data.get("make", "").strip()
    model = data.get("model", "").strip()
    year = str(data.get("year", "")).strip()
    trim = data.get("trim", "").strip()
    color = data.get("color", "").strip().lower()    # the catalog expects a lowercase basic colour
    nocache = bool(data.get("nocache"))              # "get new batch" -> skip the cache, re-query
    if not (make and model):
        return jsonify({"error": "Pick a make and model first."}), 400

    # The model now carries the powertrain for display (e.g. "Sorento Hybrid"), but the
    # image catalog only knows the base model — strip it for the lookup so a hybrid still
    # finds photos.
    q_model = _carsxe_base_model(model)

    key = "|".join([year, make.lower(), q_model.lower(), trim.lower(), color])
    cached = _CARSXE_IMG_CACHE.get(key)
    if cached is not None and not nocache:
        for o in cached.get("options", []):           # keep proxy allow-list warm
            _carsxe_remember(o["url"]); _carsxe_remember(o.get("thumb"))
        return jsonify({**cached, "cached": True})

    base = {"key": CARSXE_API_KEY, "make": make, "model": q_model,
            "transparent": "true", "size": "Large", "format": "json"}
    if year:
        base["year"] = year
    if trim:
        base["trim"] = trim
    if color:
        base["color"] = color

    # Best-first attempts: ask for exterior-only studio shots (needs year+trim),
    # then fall back to the unfiltered set so we always return something.
    attempts = []
    if year and trim:
        attempts.append({**base, "photoType": "exterior"})
    attempts.append(base)

    imgs, body = [], {}
    try:
        for p in attempts:
            url = "https://api.carsxe.com/images?" + urllib.parse.urlencode(p)
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=30) as r:
                body = json.loads(r.read().decode("utf-8", "ignore"))
            imgs = [im for im in (body.get("images") or []) if im.get("link")]
            if imgs:
                break
    except urllib.error.HTTPError as e:
        return jsonify({"error": f"Catalog error {e.code}: {e.read().decode('utf-8', 'ignore')[:200]}"}), 502
    except Exception as e:
        return jsonify({"error": f"Catalog request failed: {e}"}), 502

    if not imgs:
        return jsonify({"found": False, "message": body.get("error") or "No images returned by the catalog."})
    for im in imgs:
        _carsxe_remember(im["link"]); _carsxe_remember(im.get("thumbnailLink"))
    options = [{"url": im["link"], "thumb": im.get("thumbnailLink") or im["link"],
                "w": im.get("width"), "h": im.get("height"),
                "transparent": str(im.get("mime", "")).endswith("png")}
               for im in imgs]
    # load the first image that actually downloads (a single slow one shouldn't kill the request);
    # whichever loads becomes image_data, and we float it to the front of the options.
    first, first_url, err = None, None, None
    for im in imgs[:3]:
        try:
            first = _carsxe_dataurl(im["link"]); first_url = im["link"]; break
        except Exception as e:
            err = e
    if first is None:
        return jsonify({"error": f"Couldn't load a catalog image (the photo CDN was slow) — try again: {err}"}), 502
    if first_url and options and options[0]["url"] != first_url:
        options.sort(key=lambda o: 0 if o["url"] == first_url else 1)
    payload = {"found": True, "success": True, "image_data": first,
               "image_url": first_url, "options": options}
    _cache_put(_CARSXE_IMG_CACHE, key, payload)
    return jsonify(payload)


@app.route("/carsxe-proxy", methods=["POST"])
def carsxe_proxy():
    """Proxy a specific catalog image (only catalog/vendor-CDN URLs — no open SSRF)."""
    url = (request.json or {}).get("url", "").strip()
    if not _carsxe_proxy_ok(url):
        return jsonify({"error": "URL not allowed."}), 400
    cached = _CARSXE_PROXY_CACHE.get(url)
    if cached is not None:
        return jsonify({"success": True, "image_data": cached, "cached": True})
    try:
        d = _carsxe_dataurl(url)
        _cache_put(_CARSXE_PROXY_CACHE, url, d)
        return jsonify({"success": True, "image_data": d})
    except Exception as e:
        return jsonify({"error": str(e)}), 502


def _carvector_get(path):
    req = urllib.request.Request("https://api.carvector.io" + path,
                                 headers={"Authorization": f"Bearer {CARVECTOR_API_KEY}", "User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode("utf-8", "ignore"))


@app.route("/carvector-image", methods=["POST"])
def carvector_image():
    """TRIAL: CarVector returns one illustration per vehicle via a 2-step flow
    (search year/make/model -> get id -> fetch detail -> image_url). We gather a
    few matching rows (trims/submodels) so the user has a small set to pick from."""
    if not CARVECTOR_API_KEY:
        return jsonify({"error": "CARVECTOR_API_KEY is not set on the server. Add it in Railway → Variables."}), 500
    data = request.json or {}
    make = data.get("make", "").strip()
    model = data.get("model", "").strip()
    year = str(data.get("year", "")).strip()
    if not (make and model):
        return jsonify({"error": "Pick a make and model first."}), 400

    key = "|".join([year, make.lower(), model.lower()])
    cached = _CARVECTOR_CACHE.get(key)
    if cached is not None:
        _CARVECTOR_ALLOWED.update(o["url"] for o in cached.get("options", []))
        return jsonify({**cached, "cached": True})

    # try exact (year+make+model), then drop the year (specs DBs lag on new years),
    # then make+model only — first non-empty wins. Keep diagnostics for the UI.
    attempts = []
    if year:
        attempts.append({"make": make, "model": model, "year": year, "limit": "8"})
    attempts.append({"make": make, "model": model, "limit": "8"})

    results, tried = [], []
    try:
        for qs in attempts:
            search = _carvector_get("/v1/vehicles?" + urllib.parse.urlencode(qs))
            rs = search.get("results") or search.get("data") or search.get("vehicles") or []
            tried.append({"q": {k: v for k, v in qs.items() if k != "limit"}, "count": len(rs)})
            if rs:
                results = rs
                break
    except urllib.error.HTTPError as e:
        return jsonify({"error": f"CarVector error {e.code}: {e.read().decode('utf-8', 'ignore')[:200]}"}), 502
    except Exception as e:
        return jsonify({"error": f"CarVector request failed: {e}"}), 502

    if not results:
        return jsonify({"found": False, "message": "No vehicle found in CarVector.", "tried": tried})

    options, seen = [], set()
    for v in results[:5]:                       # cap detail calls (each one is a request)
        vid = v.get("id")
        if not vid:
            continue
        try:
            detail = _carvector_get("/v1/vehicles/" + urllib.parse.quote(str(vid)))
        except Exception:
            continue
        img = detail.get("image_url")
        if img and img not in seen:
            seen.add(img)
            label = " ".join(str(x) for x in [detail.get("year"), detail.get("trim") or detail.get("submodel") or detail.get("body_class")] if x)
            options.append({"url": img, "thumb": img, "label": label or "Illustration"})
    if not options:
        return jsonify({"found": False, "message": "CarVector has no image for that vehicle (image plans only)."})

    _CARVECTOR_ALLOWED.update(o["url"] for o in options)
    try:
        first = _carsxe_dataurl(options[0]["url"])
    except Exception as e:
        return jsonify({"error": f"Couldn't load the CarVector image: {e}"}), 502
    payload = {"found": True, "image_data": first, "image_url": options[0]["url"], "options": options}
    _cache_put(_CARVECTOR_CACHE, key, payload)
    return jsonify(payload)


@app.route("/carvector-proxy", methods=["POST"])
def carvector_proxy():
    url = (request.json or {}).get("url", "").strip()
    if url not in _CARVECTOR_ALLOWED:
        return jsonify({"error": "URL not allowed."}), 400
    try:
        return jsonify({"success": True, "image_data": _carsxe_dataurl(url)})
    except Exception as e:
        return jsonify({"error": str(e)}), 502


@app.route("/bulk-parse", methods=["POST"])
def bulk_parse():
    """Smart bulk: an LLM turns ANY pasted text (list, email, spreadsheet dump)
    into structured ad rows the user can review and download."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    text = (request.json or {}).get("text", "").strip()
    if not text:
        return jsonify({"error": "Paste some vehicles first."}), 400
    schema = {
        "type": "object", "additionalProperties": False, "required": ["rows"],
        "properties": {"rows": {"type": "array", "items": {
            "type": "object", "additionalProperties": False,
            "required": ["year", "make", "model", "trim", "deal_type", "monthly", "das", "down", "apr", "term", "badge"],
            "properties": {
                "year": {"type": "string"}, "make": {"type": "string"}, "model": {"type": "string"},
                "trim": {"type": "string"},
                "deal_type": {"type": "string", "enum": ["lease", "finance", "onepay"]},
                "monthly": {"type": "string"}, "das": {"type": "string"}, "down": {"type": "string"},
                "apr": {"type": "string"}, "term": {"type": "string"}, "badge": {"type": "string"},
            }}}},
    }
    system = (
        "You extract vehicle lease/finance offers from messy pasted text into ad rows. "
        "For each vehicle fill: year; make (FULL make, e.g. 'Mercedes-Benz', 'BMW'); model; trim; "
        "deal_type (lease | finance | onepay); monthly (monthly payment); das (due-at-signing, lease); "
        "down (down payment, finance); apr (finance APR percent); term (months); badge (promo label if any). "
        "POWERTRAIN ('Hybrid', 'PHEV', 'Plug-in', 'Electric', 'EV') is part of the MODEL, not the trim — "
        "append it to the model so it shows on the ad (e.g. 'Kia Sorento Hybrid LX' -> model 'Sorento Hybrid', "
        "trim 'LX'; 'RAV4 Hybrid XLE' -> model 'RAV4 Hybrid', trim 'XLE'). Omit plain 'Gas' (it's the default). "
        "Use '' for anything not present. Numbers only — no $, no commas. If it says one-pay/single-pay, "
        "deal_type=onepay and put the up-front total in das. Default deal_type=lease when unclear."
    )
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "system", "content": system}, {"role": "user", "content": text}],
            response_format={"type": "json_schema", "json_schema": {"name": "bulk", "strict": True, "schema": schema}},
        )
        rows = json.loads(resp.choices[0].message.content).get("rows", [])
        return jsonify({"rows": rows})
    except Exception as e:
        return jsonify({"error": f"Couldn't parse that: {e}"}), 502


# ----------------------- Deal Hub -----------------------
DEALS_FILE = os.path.join(GENERATED_DIR, "deals.json")
DEAL_FIELDS = ["year", "make", "model", "trim", "package", "msrp", "orig_mo", "das", "term",
               "miles", "tax_in_mo", "broker_fee", "dealer", "notes", "source",
               "active_from", "active_to", "special", "deal_type", "apr", "incentives"]


def _load_deals():
    try:
        with open(DEALS_FILE) as f:
            d = json.load(f)
            return d if isinstance(d, list) else []
    except Exception:
        return []


def _save_deals(deals):
    with open(DEALS_FILE, "w") as f:
        json.dump(deals, f, indent=1)


def _default_year(d):
    """Most of our cars are new leases — if a year isn't stated, assume the current model year."""
    if not str(d.get("year") or "").strip():
        d["year"] = "2026"
    return d


def _strip_tax(d):
    """We NEVER keep tax-included figures. Any tax_in_mo is backed out to a pre-tax
    orig_mo (÷1.0975) and cleared — a hard guarantee on top of the model's own instruction."""
    try:
        t = d.get("tax_in_mo")
        if t not in (None, "", "0"):
            tv = float(str(t).replace("$", "").replace(",", "").strip())
            if tv:
                if not d.get("orig_mo"):
                    d["orig_mo"] = str(round(tv / 1.0975, 2))
                d["tax_in_mo"] = ""
    except Exception:
        pass
    return d


def _derive_zero_down(d):
    """0-Down Mo = Orig Mo + DAS / Term (the dealer-sheet convention)."""
    try:
        mo = float(d.get("orig_mo") or 0)
        das = float(d.get("das") or 0)
        term = float(d.get("term") or 0)
        if mo and term:
            d["zero_down_mo"] = str(round(mo + das / term, 2))
    except Exception:
        pass
    return d


@app.route("/deal-parse", methods=["POST"])
def deal_parse():
    """Read messy pasted deal text into structured rows, applying the user's
    plain-English adjustment rules (math) and propagating global header terms."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    body = request.json or {}
    text = (body.get("text") or "").strip()
    # rules used to be a separate field; now the AI also picks adjustments out of the
    # paste itself. Still honored if a caller passes it explicitly.
    rules = (body.get("rules") or "").strip()
    makes = [m for m in (body.get("makes") or []) if isinstance(m, str) and m.strip()][:80]
    today = (body.get("today") or "").strip()
    images = [im for im in (body.get("images") or []) if isinstance(im, str) and im.startswith("data:image")][:8]
    if not text and not images:
        return jsonify({"error": "Paste some deals or attach a flyer first."}), 400
    date_rule = (
        f"- Today is {today}. If a line/list states a validity window ('good through 5/31', "
        "'May specials', 'valid until 6/15', 'expires EOM'), output active_from / active_to as "
        "YYYY-MM-DD. 'May specials' => first..last day of that May. If only an end is given, set "
        "active_to and leave active_from ''. Leave BOTH '' if no dates are mentioned.\n"
    ) if today else ""
    make_rule = (
        "- NORMALIZE 'make' to EXACTLY one entry (matching case) from this canonical list: "
        + ", ".join(makes) + ". Map shorthand/synonyms to it: 'Chevy'->'Chevrolet', "
        "'VW'->'Volkswagen', 'Mercedes'/'Benz'/'MB'->'Mercedes-Benz', 'Range Rover'->'Land Rover', "
        "'Alfa'->'Alfa Romeo'. If a make truly isn't in the list, output your best clean guess.\n"
        "- Clean the 'model' to its standard spelling/casing ('k5'->'K5', 'corolla cross'->'Corolla Cross').\n"
        "- Use the manufacturer's STANDARD spacing for alphanumeric model codes and be CONSISTENT — "
        "always write a given model the same way (e.g. 'IS 350', 'RX 350', 'GLC 300', 'F-150'); never "
        "emit the same vehicle two different ways ('IS 350' on one line, 'IS350' on the next).\n"
    ) if makes else ""
    props = {k: {"type": "string"} for k in DEAL_FIELDS}
    schema = {
        "type": "object", "additionalProperties": False,
        "required": ["kind", "contact", "contact_phone", "contact_email", "deals", "adjust"],
        "properties": {
            "kind": {"type": "string", "enum": ["deals", "adjust"]},
            "contact": {"type": "string"},
            "contact_phone": {"type": "string"},
            "contact_email": {"type": "string"},
            "deals": {"type": "array", "items": {
                "type": "object", "additionalProperties": False,
                "required": DEAL_FIELDS, "properties": props}},
            "adjust": {
                "type": "object", "additionalProperties": False,
                "required": ["match", "ops", "remove_tax_pct"],
                "properties": {
                    "match": {
                        "type": "object", "additionalProperties": False,
                        "required": ["contact", "year", "make", "model", "trim"],
                        "properties": {k: {"type": "string"} for k in ["contact", "year", "make", "model", "trim"]}},
                    "ops": {
                        "type": "array",
                        "items": {
                            "type": "object", "additionalProperties": False,
                            "required": ["field", "op", "value"],
                            "properties": {
                                "field": {"type": "string", "enum": ["monthly", "das", "broker_fee", "term", "miles", "msrp", "apr"]},
                                "op": {"type": "string", "enum": ["add", "mult", "set"]},
                                "value": {"type": "string"}}}},
                    "remove_tax_pct": {"type": "string"}}},
        },
    }
    system = (
        "You read messy car lease/finance text. FIRST classify it with 'kind':\n"
        "- 'adjust' = an INSTRUCTION to change deals ALREADY saved, with NO new vehicle listings "
        "(e.g. 'decrease all GLA 250 das by $1000', 'lower Diana Santa Monica's monthlies $15', "
        "'Mike: 9.75% tax included'). Output an 'adjust' object and leave 'deals' empty.\n"
        "- 'deals' = an actual list/flyer of one or more vehicles. Output 'deals' and leave "
        "'adjust' empty (match all '', ops []).\n"
        "ADJUST object: 'match' selects WHICH saved deals to change — set any of contact / year / "
        "make / model / trim that the instruction names, leave the rest '' (all-'' = EVERY deal). "
        "'ops' is a list of {field, op, value}: field is monthly|das|broker_fee|term|miles|msrp|apr "
        "(use 'monthly' for the payment, 'das' for down/DAS, 'apr' for finance rate); op is 'add' "
        "(value SIGNED, '-1000' to decrease, '20' to "
        "add), 'mult' (multiply), or 'set' (replace). remove_tax_pct = a tax rate to back OUT of "
        "the monthly when they say tax-included ('9.75% tax included'->'9.75'; plain 'tax included' "
        "with no rate -> '9.75'; else '').\n"
        "  A bare brand name is a MAKE, not a model: 'update all Toyota deals ...' -> match:{make:'Toyota'} (NOT model).\n"
        "  e.g. 'decrease all GLA 250 das by $1000' -> match:{model:'GLA 250'}, ops:[{field:'das',op:'add',value:'-1000'}].\n"
        "  e.g. 'add $500 to every Toyota broker fee' -> match:{make:'Toyota'}, ops:[{field:'broker_fee',op:'add',value:'500'}].\n"
        "  e.g. 'lower Diana's monthlies $15' -> match:{contact:'Diana'}, ops:[{field:'monthly',op:'add',value:'-15'}].\n"
        "For EACH vehicle (kind='deals') output: year, make, model, trim, package (equipment "
        "PACKAGES / options — see rule below), msrp, orig_mo (base "
        "monthly payment), das (due at signing / cash to dealer), term (months), miles (annual "
        "mileage), tax_in_mo (monthly WITH tax, if quoted that way), broker_fee, dealer, notes "
        "(loyalty / conquest / credit tier / any other terms), source (short label for where it "
        "came from), active_from / active_to (validity window, only if stated), special (a "
        "short tag like 'Loaner', '1 of 1', or 'Demo' ONLY if the deal is a loaner / courtesy / "
        "demo / service-loaner / one-off special; else ''), deal_type ('finance' if it's a "
        "incentives (a comma-separated list of QUALIFYING rebates the customer must qualify for: "
        "'Loyalty', 'Conquest', 'AAA', 'College Grad', 'Military', 'First Responder', etc. — use the "
        "canonical name ('AAA', never 'Triple A'), never list the same one twice, and keep these OUT "
        "of notes), deal_type ('finance' if it's a "
        "FINANCE/purchase/loan deal — mentions APR, % financing, 'finance', months-to-own, down "
        "payment — otherwise 'lease'), and apr (the finance APR as a number like '4.9', only for "
        "finance deals; else '').\n"
        "  For a FINANCE deal: orig_mo = the monthly payment, das = the DOWN PAYMENT (cash up "
        "front), term = months, apr = the rate; miles is usually '' (no mileage cap on a purchase).\n"
        "ALSO output top-level 'contact': the single person/broker/dealer the paste is from "
        "(sender at top, signature, or 'from X'). If unsure, ''. Put that name in each row's "
        "'dealer' field too. Also output 'contact_phone' and 'contact_email' if a phone number or "
        "email for that contact appears anywhere in the text (else '').\n"
        "RULES:\n"
        "- If a vehicle's YEAR isn't stated, use '2026' (most of our cars are new leases).\n"
        "- PACKAGES / OPTIONS are NOT part of the car. Keep equipment packages and options "
        "('Convenience Package', 'Premium Package', 'Technology Package', 'Cold Weather Package', "
        "'Shadowline', etc.) in 'package' and OUT of model and trim. model+trim identify the actual "
        "vehicle; trim is the variant/trim level (e.g. 'M Sport', 'xDrive', 'GT-Line', 'Luxury'). "
        "e.g. '2026 BMW 330i Convenience Package' -> make 'BMW', model '330i', trim '', package "
        "'Convenience Package'.\n"
        "- POWERTRAIN ('Hybrid', 'PHEV', 'Plug-in', 'Plug-in Hybrid', 'Electric', 'EV') is part of "
        "the MODEL, not the trim or package — append it to the model so it shows on the ad (clients "
        "must see it's a hybrid). e.g. 'Kia Sorento Hybrid LX' -> model 'Sorento Hybrid', trim 'LX'; "
        "'RAV4 Hybrid XLE' -> model 'RAV4 Hybrid', trim 'XLE'; 'NX 450h+ Luxury' -> model 'NX 450h+', "
        "trim 'Luxury'. OMIT a plain 'Gas' entirely (it's the default): 'RX 350 Gas Premium' -> model "
        "'RX 350', trim 'Premium'. Never put a powertrain word in 'package'.\n"
        "- EXPAND abbreviated trim levels to the full standard name and stay CONSISTENT: 'PREM'->"
        "'Premium', 'Prem Plus'/'PREM+'->'Premium Plus', 'LUX'->'Luxury', 'Ult Lux'->'Ultra Luxury', "
        "'Ltd'->'Limited', 'Plat'->'Platinum', 'Pref'->'Preferred', 'F-Sport'->'F Sport'.\n"
        + make_rule + date_rule +
        "- GLOBAL header terms (e.g. '$2000 Down, 36 Month, 10k Miles, $1000 BF') apply to EVERY "
        "line below — copy them into each row.\n"
        "- ONE car with MULTIPLE cash-down options: if a SINGLE vehicle is quoted as a matrix of "
        "cash-down -> monthly choices at the SAME term/miles (e.g. '$3,500 down=$820/mo, $4,500=$773, "
        "$5,500=$734', often shown side by side), output ONLY ONE deal for that car — the option with "
        "the LOWEST monthly payment (usually the highest cash down). Put that monthly in orig_mo and "
        "its matching cash-down in das. Do NOT create a separate deal per down option (we let the user "
        "re-slide the down on our side).\n"
        "- Different TERMS or different annual MILES for the SAME car ARE separate offers — output a "
        "SEPARATE deal for EACH (e.g. one car shown as '39mo 7500 due $2197/m' AND '24mo 7500 due "
        "$2412/m' = TWO deals). Only the cash-down matrix above (same term, different downs) collapses "
        "to one.\n"
        "- Inline ADJUSTMENT instructions mixed into a deal list ('subtract $15 from each "
        "monthly', 'add my $500 fee') — do the arithmetic on every row; never treat as a vehicle.\n"
        "- PER-IMAGE instructions: attached images are labeled 'Screenshot 1', 'Screenshot 2', … in "
        "order. The text may target them individually ('for the first screenshot add $40 to each "
        "monthly', 'screenshot 2: add a $1000 broker fee', 'second one is from Roji'). Apply each "
        "such instruction ONLY to the deals from that screenshot.\n"
        "- TAX — two OPPOSITE cases, never confuse them:\n"
        "  (a) PLUS tax ('475+ tax', '$475 plus tax', '475 + tax', '475 +tax'): the number is "
        "ALREADY the PRE-TAX payment (tax is added on top, separately). KEEP it as-is in orig_mo. "
        "Do NOT divide.\n"
        "  (b) TAX-INCLUDED ('tax in', 'tax included', 'w/ tax', 'with tax', 'incl tax', 'taxes in'): "
        "the number INCLUDES tax. Back it out — divide by 1.0975 (our 9.75%; or 1+rate/100 if a "
        "different rate is given) — and put the PRE-TAX result in orig_mo.\n"
        "  Either way leave tax_in_mo '' and never mention tax in notes. (We NEVER store a "
        "tax-included number; '+ tax' is NOT tax-included.)\n"
        "- Normalize shorthand to plain numbers (no $ or commas): '3k'->3000, '7.5k'->7500, "
        "'$51k MSRP'->51000, '$289'->289. '13/7500' means term=13, miles=7500.\n"
        "- Leave a field '' if it isn't present.\n"
        f"EXTRA ADJUSTMENT RULES (if any) TO APPLY TO EVERY ROW: {rules or '(none)'}"
    )
    if images:
        # vision: read the deals straight off a flyer screenshot / PDF page(s)
        model = "gpt-4.1"
        user_content = [{"type": "text", "text": text or "Read EVERY car deal from these flyer image(s) and structure them. Capture each vehicle, price, term, miles/down, dealer/contact, and any dates."}]
        for n, im in enumerate(images, 1):
            user_content.append({"type": "text", "text": f"--- Screenshot {n} ---"})
            user_content.append({"type": "image_url", "image_url": {"url": im}})
    else:
        model = "gpt-4.1-mini"
        user_content = text
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "system", "content": system}, {"role": "user", "content": user_content}],
            response_format={"type": "json_schema", "json_schema": {"name": "deals", "strict": True, "schema": schema}},
        )
        out = json.loads(resp.choices[0].message.content)
        deals = [_derive_zero_down(_strip_tax(_default_year(d))) for d in out.get("deals", [])]
        return jsonify({
            "kind": out.get("kind") or "deals",
            "deals": deals,
            "contact": (out.get("contact") or "").strip(),
            "contact_phone": (out.get("contact_phone") or "").strip(),
            "contact_email": (out.get("contact_email") or "").strip(),
            "adjust": out.get("adjust") or {},
        })
    except Exception as e:
        return jsonify({"error": f"Couldn't parse those deals: {e}"}), 502


@app.route("/desk-parse", methods=["POST"])
def desk_parse():
    """Read a dealer lease/finance RATE SHEET (pasted text or a screenshot) into the
    structured program inputs the Desking calculator needs: MSRP, residual, money
    factor / APR, term, miles, acquisition fee, rebates, fees. Picks ONE primary
    program when several term/mileage rows are listed."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    body = request.json or {}
    text = (body.get("text") or "").strip()
    images = [im for im in (body.get("images") or []) if isinstance(im, str) and im.startswith("data:image")][:6]
    if not text and not images:
        return jsonify({"error": "Paste a rate sheet or attach a screenshot first."}), 400
    fields = ["vehicle", "deal_type", "msrp", "selling_price", "residual_pct",
              "residual_amount", "money_factor", "apr", "term", "miles",
              "acq_fee", "rebate", "fees", "down", "tax_pct", "notes"]
    term_keys = ["term", "residual_pct", "money_factor", "miles"]
    schema = {
        "type": "object", "additionalProperties": False,
        "required": fields + ["terms"],
        "properties": {
            **{k: {"type": "string"} for k in fields if k != "deal_type"},
            "deal_type": {"type": "string", "enum": ["lease", "finance"]},
            "terms": {
                "type": "array",
                "items": {
                    "type": "object", "additionalProperties": False,
                    "required": term_keys,
                    "properties": {k: {"type": "string"} for k in term_keys},
                },
            },
        },
    }
    system = (
        "You read a car dealer's LEASE/FINANCE RATE SHEET or program bulletin (the monthly "
        "captive-lender numbers a dealer uses to structure a deal) and extract the inputs a "
        "desking calculator needs. Output ONE program object.\n"
        "FIELDS:\n"
        "- vehicle: 'Year Make Model Trim' if present (e.g. '2026 Toyota RAV4 XLE'); else as much as given.\n"
        "- deal_type: 'lease' if it quotes residual/money factor/MF/lease rate; 'finance' if it quotes "
        "APR/financing/purchase. Default 'lease' when ambiguous but a residual or money factor appears.\n"
        "- msrp: the MSRP / sticker (plain number, no $ or commas).\n"
        "- selling_price: the selling/sale/cap-cost/dealer price ONLY if explicitly given. Rate sheets "
        "usually DON'T have it — leave '' if absent (do NOT guess it from MSRP).\n"
        "- residual_pct: the residual as a percent number ('60', '62.5'). If the sheet gives a residual "
        "DOLLAR amount instead, put that in residual_amount and leave residual_pct '' (we'll compute it).\n"
        "- residual_amount: residual VALUE in dollars if given as a dollar figure; else ''.\n"
        "- money_factor: the lease money factor as a decimal ('0.00150'). If only a lease RATE / lease "
        "APR percent is given, CONVERT it to a money factor by dividing by 2400 ('3.6%' -> '0.00150') "
        "and output that. Else ''.\n"
        "- apr: finance APR as a number ('5.9') for finance deals; else ''.\n"
        "- term: months ('36'). If several terms are listed, PICK 36 if present, otherwise the most "
        "standard / first one.\n"
        "- miles: annual mileage for a lease ('10000'); if several, match the term you picked, preferring "
        "10000 then 12000 then 7500; else ''.\n"
        "- acq_fee: acquisition / bank fee ('650'); else ''.\n"
        "- rebate: total customer rebates / incentives / lease cash as a number; else ''.\n"
        "- fees: doc / other dealer fees as a number; else ''.\n"
        "- down: stated cap-cost-reduction / down payment if any; else ''.\n"
        "- tax_pct: sales-tax rate ONLY if the sheet states one; else '' (caller defaults 9.75).\n"
        "- notes: anything important that doesn't fit (credit tier, region/zone, validity dates, "
        "loyalty/conquest requirement, multiple-term summary). Keep it short; else ''.\n"
        "- terms: if the sheet lists MULTIPLE TERMS (e.g. 24 / 36 / 39 / 48 month) each with their OWN "
        "residual and money factor, output one row PER TERM in this array: {term, residual_pct, "
        "money_factor, miles}. Use the same conversions (residual as %, money factor as a small decimal, "
        "lease-rate% ÷ 2400 -> MF). Use the PRIMARY mileage row for each term (prefer 10k, then 12k, then "
        "7.5k). If a term residual is given only in dollars, convert to % using MSRP. If the sheet shows "
        "only one term (no real curve), output terms = [] (empty). Still fill the single residual_pct / "
        "money_factor / term fields above with the primary (36-month) row.\n"
        "RULES: normalize shorthand to plain numbers ('51k'->51000, '$3,500'->3500, '.0012 MF'->0.0012). "
        "A money factor is a small decimal like 0.00120; never confuse it with APR percent. Leave any "
        "field '' if it isn't present. Output numbers only (no $, %, commas) except 'vehicle' and 'notes'."
    )
    if images:
        model = "gpt-4.1"
        user_content = [{"type": "text", "text": text or "Read the lease/finance program off this rate sheet and extract the desking inputs."}]
        for n, im in enumerate(images, 1):
            user_content.append({"type": "text", "text": f"--- Sheet {n} ---"})
            user_content.append({"type": "image_url", "image_url": {"url": im}})
    else:
        model = "gpt-4.1-mini"
        user_content = text
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "system", "content": system}, {"role": "user", "content": user_content}],
            response_format={"type": "json_schema", "json_schema": {"name": "program", "strict": True, "schema": schema}},
        )
        out = json.loads(resp.choices[0].message.content)
        prog = {k: (out.get(k) or "").strip() if isinstance(out.get(k), str) else out.get(k) for k in fields}
        rows = out.get("terms") if isinstance(out.get("terms"), list) else []
        prog["terms"] = [
            {"term": (r.get("term") or "").strip(), "resid": (r.get("residual_pct") or "").strip(),
             "mf": (r.get("money_factor") or "").strip(), "miles": (r.get("miles") or "").strip()}
            for r in rows if isinstance(r, dict) and (r.get("term") or "").strip()
        ]
        return jsonify({"program": prog})
    except Exception as e:
        return jsonify({"error": f"Couldn't read that rate sheet: {e}"}), 502


LIVE_DEALS_URL = "https://novautousa.com/deals"
_LIVE_DEALS_CACHE = {"at": 0.0, "text": "", "count": 0}

# ---- Concierge Copilot knowledge base (the NovAuto sales guide, etc.) ----
# Two private sources, NEITHER in the public repo:
#   1) COPILOT_KB env var on Railway  ← set it once, no in-app upload needed (preferred)
#   2) the volume file written by the in-app 📚 editor (overrides the env var if used)
COPILOT_KB_FILE = os.path.join(GENERATED_DIR, "copilot_kb.txt")
_COPILOT_KB_CACHE = {"at": 0.0, "text": None}


def _load_copilot_kb(max_age=60):
    now = time.time()
    if _COPILOT_KB_CACHE["text"] is not None and (now - _COPILOT_KB_CACHE["at"] < max_age):
        return _COPILOT_KB_CACHE["text"]
    text = ""
    try:                                                  # in-app editor file wins if present
        with open(COPILOT_KB_FILE, encoding="utf-8") as f:
            text = f.read().strip()
    except Exception:
        text = ""
    if not text:                                          # else the Railway env var
        text = (os.environ.get("COPILOT_KB") or "").strip()
    _COPILOT_KB_CACHE.update(at=now, text=text)
    return text


@app.route("/copilot-kb", methods=["GET", "POST"])
def copilot_kb():
    """Admin (restricted) read/write of the copilot knowledge base. The text lives on
    the volume only — never in the public repo. Injected into the copilot's prompt."""
    if request.method == "POST":
        text = ((request.json or {}).get("text") or "").strip()
        try:
            with open(COPILOT_KB_FILE, "w", encoding="utf-8") as f:
                f.write(text)
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        _COPILOT_KB_CACHE.update(at=0.0, text=None)        # bust cache
        return jsonify({"ok": True, "chars": len(text)})
    return jsonify({"text": _load_copilot_kb(max_age=0)})


def _fetch_live_deals_text(max_age=300):
    """Fetch & cache the live novautousa.com/deals as a compact text block the Broker
    Copilot can read. Cached ~5 min; on failure, returns any stale cache. -> (text, count)."""
    now = time.time()
    if _LIVE_DEALS_CACHE["text"] and (now - _LIVE_DEALS_CACHE["at"] < max_age):
        return _LIVE_DEALS_CACHE["text"], _LIVE_DEALS_CACHE["count"]
    try:
        req = urllib.request.Request(LIVE_DEALS_URL, headers={"User-Agent": "Mozilla/5.0 (NovaCopilot)"})
        with urllib.request.urlopen(req, timeout=15) as r:
            body = r.read(3_000_000).decode("utf-8", "ignore")
        rows = _extract_inertia_deals(body) or []
        lines = []
        for d in rows:
            veh = " ".join(x for x in [d.get("year", ""), d.get("make", ""), d.get("model", ""), d.get("trim", "")] if x).strip()
            parts = [veh] if veh else []
            if d.get("monthly"):
                parts.append("$" + str(d["monthly"]) + "/mo")
            if d.get("das"):
                parts.append("$" + str(d["das"]) + " due at signing")
            if d.get("term"):
                parts.append(str(d["term"]) + "mo")
            if d.get("slug"):
                parts.append("page: https://novautousa.com/deals/" + urllib.parse.quote(d["slug"], safe=""))
            if parts:
                lines.append("• " + " · ".join(parts))
        text = "\n".join(lines)
        if text:
            _LIVE_DEALS_CACHE.update(at=now, text=text, count=len(lines))
        return text, len(lines)
    except Exception:
        return _LIVE_DEALS_CACHE["text"], _LIVE_DEALS_CACHE["count"]


@app.route("/broker-chat", methods=["POST"])
def broker_chat():
    """Concierge Copilot — a conversational assistant for NovAuto's agents, available
    on every page. Multi-turn; the client sends the running message history. Login-gated
    (every route is) but NOT behind the restricted password, so any agent can use it."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    body = request.json or {}
    raw = body.get("messages") or []
    clean = []
    for m in raw[-20:]:
        if not isinstance(m, dict):
            continue
        role = m.get("role")
        content = (m.get("content") or "").strip()
        if role in ("user", "assistant") and content:
            clean.append({"role": role, "content": content[:4000]})
    if not clean or clean[-1]["role"] != "user":
        return jsonify({"error": "Ask a question first."}), 400
    system = (
        "You are 'Concierge Copilot', the in-house AI assistant for NovAuto — a California auto "
        "LEASE & FINANCE concierge that sources and negotiates new cars for clients. "
        "You help NovAuto's concierge agents do their job. Be warm, sharp, and concise — like a seasoned "
        "finance manager mentoring a newer agent. Plain language, short paragraphs, bullets when useful. "
        "Avoid long essays; get to the point and offer to go deeper.\n"
        "You can help with: lease & finance math (money factor, residual, capitalized cost, depreciation + "
        "rent charge, due at signing, APR, amortization), explaining those to clients simply, structuring and "
        "presenting deals, objection handling and talk tracks, lease vs finance, mileage/term tradeoffs, and how "
        "to use the NovAuto toolkit — Deal Hub (live deal board), Desking (deal calculator with a down×term grid, "
        "mileage→residual, a saved program library, and AI rate-sheet parsing), Lease Ad & Sold Posts (marketing "
        "graphics), Review Generator, Invoice Generator, and Deal Proposal.\n"
        "VOICE RULES: ALWAYS call the buyer a 'client', NEVER a 'customer'. The brand motto is "
        "\"Friends Don't Let Friends Overpay.\" Professional but friendly.\n"
        "MATH YOU KNOW: monthly money factor ≈ APR ÷ 2400. Lease monthly = depreciation [(adjusted cap − "
        "residual) ÷ term] + rent charge [(adjusted cap + residual) × money factor], then add CA sales tax on the "
        "payment. Residual is a % of MSRP. More miles/year → lower residual → higher payment. Down payment / "
        "rebates reduce the cap cost. Due at signing = down + first payment + broker fee. Finance: tax the selling "
        "price, finance it, standard amortization. Show the formula and a worked number when it helps.\n"
        "BALLPARK QUICK-REFERENCE (Nova rules of thumb): $1,000 cap reduction ≈ $30/mo less on 36mo (≈ $42/mo on "
        "24mo); $500 cap reduction ≈ $15/mo less on 36mo; every 1% residual drop ≈ $15–20/mo more; a down-payment "
        "change ÷ term = the monthly change ($1k less down ÷ 36 ≈ +$28/mo); MF × 2400 = APR; LA tax: ×1.0975 to add, "
        "÷1.0975 to remove. Every number we give a client is a TRUE OUT-THE-DOOR number — tax, DMV, fees, first "
        "payment, and the Nova concierge fee all included; quote ONE clean DAS and ONE clean monthly.\n"
        "BALLPARK ESTIMATES: If an agent asks roughly what a car leases/finances for and it's NOT pre-negotiated "
        "(not in our live deals), give a useful BALLPARK by (a) using similar cars from our live deals as comps, (b) "
        "applying the rules of thumb above, and (c) your general market knowledge for that segment. ALWAYS label it "
        "as a 'rough ballpark, not a locked quote', state your assumptions (term / miles / DAS), and tell them to "
        "confirm the real number in Desking or by sourcing through Nema. Give a sensible monthly RANGE, not false "
        "precision. Never present an estimate as a firm price.\n"
        "LIVE DEALS: You DO have real-time read access to the deals published on novautousa.com/deals — they're "
        "given to you below. Use them whenever an agent asks what's live / advertised / the best or cheapest deal / "
        "a specific car / what's under a price. Quote the exact advertised monthly, due-at-signing, and term, and "
        "say it's the current advertised number. If a car they ask about isn't in the live list, say it's not "
        "currently advertised on the site.\n"
        "DEAL PAGE LINKS: each live deal below includes its 'page:' link on novautousa.com. When you reference or "
        "recommend a specific live car, ALWAYS include that page link so the agent can send it to the client — this "
        "is Nova's own marketing page (NOT the dealer's listing), so it's safe to share. (The guide's 'never send "
        "the car link' rule is about hiding the DEALER's listing/VIN, not Nova's own site.) Use the exact URL given; "
        "never invent a link.\n"
        "LIMITS: You do NOT know raw manufacturer programs (residual %, money factor, incentives) or a specific "
        "client's file — for those, point them to the Desking program library or the Deal Hub. Never INVENT a "
        "residual %, money factor, or incentive amount. Ask a brief clarifying question if a request is ambiguous. "
        "Keep general guidance only on legal/tax matters.\n"
        "ACTION BUTTONS: when your answer naturally leads the agent to USE a tool, end your message with ONE final "
        "line in EXACTLY this format and nothing after it: [[ACTIONS: key1, key2]] — using only these keys: "
        "desk (open the Desking calculator), quote (build a Deal Proposal), invoice (build an Invoice), deals (open "
        "the Deal Hub), ads (make a lease ad), sold (make a sold post), review (make a review post). Include 1–3 of "
        "the MOST relevant keys (e.g. someone closing a deal → 'quote, invoice'; pricing a car → 'desk'). If no tool "
        "is relevant, OMIT the line entirely. Never mention or explain this line; just append it when useful.\n"
        "DEAL HAND-OFF: when you include a quote/invoice/desk action AND a specific vehicle or deal is in play, ALSO "
        "append ONE more final line in EXACTLY this format (after the ACTIONS line): [[DEAL: {json}]] — a compact "
        "JSON object with any of these STRING fields you actually know from the conversation or live deals: vehicle "
        "('2026 Toyota RAV4 XLE'), dealType ('lease' or 'finance'), monthly, term, das (due at signing or down), "
        "miles, msrp, apr, color, vin, client (the client's first name), notes. Use live-deal numbers for an "
        "advertised car, or your ballpark for an estimate. OMIT fields you don't know — never invent a VIN or MSRP. "
        "This pre-fills the tool for the agent. If no specific car is in play, omit the DEAL line. Never explain it."
    )
    kb = _load_copilot_kb()
    if kb:
        system += ("\n\n=== NOVAUTO PLAYBOOK / SALES GUIDE (internal knowledge — treat as authoritative; "
                   "follow its process, terminology, scripts, and policies when answering) ===\n" + kb[:16000])
    deals_text, deals_n = _fetch_live_deals_text()
    if deals_text:
        system += ("\n\n=== LIVE DEALS on novautousa.com/deals right now (" + str(deals_n) +
                   " vehicles, refreshed within ~5 min) — these are the current client-facing advertised "
                   "numbers ===\n" + deals_text[:7000])
    else:
        system += ("\n\n(Live deals from novautousa.com/deals are momentarily unavailable — if asked about live "
                   "deals, say so and suggest checking the Deal Hub.)")
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "system", "content": system}] + clean,
            temperature=0.5, max_tokens=700,
        )
        reply = (resp.choices[0].message.content or "").strip()
        actions, deal = [], {}
        allowed = {"desk", "quote", "proposal", "invoice", "deals", "ads", "sold", "review"}
        deal_fields = {"vehicle", "dealType", "monthly", "term", "das", "miles", "msrp", "apr", "color", "vin", "client", "notes"}
        # optional [[DEAL: {json}]] → pre-fill payload for the tools
        dm = re.search(r'\[\[\s*DEAL\s*:\s*(\{.*?\})\s*\]\]', reply, re.I | re.S)
        if dm:
            try:
                d = json.loads(dm.group(1))
                if isinstance(d, dict):
                    deal = {k: str(v).strip()[:120] for k, v in d.items()
                            if k in deal_fields and v not in (None, "") and str(v).strip()}
            except Exception:
                deal = {}
        # optional [[ACTIONS: ...]] → smart buttons
        am = re.search(r'\[\[\s*ACTIONS?\s*:\s*([^\]]+)\]\]', reply, re.I)
        if am:
            for k in re.split(r'[,\s]+', am.group(1).strip().lower()):
                if k in allowed and k not in actions:
                    actions.append(k)
        starts = [x.start() for x in (am, dm) if x]
        if starts:
            reply = reply[:min(starts)].rstrip()          # strip the markers from the visible text
        return jsonify({"reply": reply, "actions": actions, "deal": deal})
    except Exception as e:
        return jsonify({"error": f"Couldn't reach the assistant: {e}"}), 502


@app.route("/deal-search", methods=["POST"])
def deal_search():
    """Natural-language filter over the deal list. Given a query and the candidate
    vehicles, an LLM returns the indices that match, best-first."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    body = request.json or {}
    query = (body.get("query") or "").strip()
    cars = body.get("cars") or []
    if not query or not isinstance(cars, list) or not cars:
        return jsonify({"match": []})
    cars = cars[:400]
    lines = []
    for i, c in enumerate(cars):
        lines.append(
            f"{i}: {c.get('year','')} {c.get('make','')} {c.get('model','')} {c.get('trim','')} "
            f"| type={c.get('type','lease')} mo={c.get('monthly','')} eff={c.get('zero','')} "
            f"down={c.get('das','')} apr={c.get('apr','')} term={c.get('term','')} miles={c.get('miles','')}"
        )
    schema = {
        "type": "object", "additionalProperties": False, "required": ["match"],
        "properties": {"match": {"type": "array", "items": {"type": "integer"}}},
    }
    system = (
        "You are a search filter for a car deal list. Given a natural-language query and a "
        "numbered list of vehicles (type lease|finance, monthly payment, effective monthly, "
        "down/DAS, apr, term months, annual miles), return 'match': the indices of vehicles that "
        "satisfy the query, MOST relevant / best first. Handle lease vs finance ('finance deals', "
        "'lease only'), APR limits ('under 5% apr'). Understand body styles (SUV, sedan, coupe, "
        "truck, wagon), fuel "
        "type (EV / hybrid) from model knowledge, brands and sub-brands ('AMG'=Mercedes-AMG, "
        "'M'/'M3'=BMW M, 'Type R'=Honda, etc.), and numeric limits ('under $500/mo' -> monthly<=500, "
        "'0 down' favors low DAS, 'cheap'/'best' -> lowest 0-down effective). Only use indices from "
        "the list. If nothing matches, return an empty array."
    )
    user = "QUERY: " + query + "\n\nVEHICLES:\n" + "\n".join(lines)
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
            response_format={"type": "json_schema", "json_schema": {"name": "search", "strict": True, "schema": schema}},
        )
        idx = json.loads(resp.choices[0].message.content).get("match", [])
        keys = [cars[i].get("key") for i in idx if isinstance(i, int) and 0 <= i < len(cars)]
        return jsonify({"keys": keys})
    except Exception as e:
        return jsonify({"error": f"Search failed: {e}"}), 502


@app.route("/deals", methods=["GET", "POST"])
def deals():
    if request.method == "POST":
        incoming = (request.json or {}).get("deals", [])
        if not isinstance(incoming, list):
            return jsonify({"error": "Bad payload."}), 400
        _save_deals(incoming)
        return jsonify({"ok": True, "count": len(incoming)})
    return jsonify({"deals": _load_deals()})


CONTACTS_FILE = os.path.join(GENERATED_DIR, "contacts.json")


@app.route("/contacts", methods=["GET", "POST"])
def contacts():
    if request.method == "POST":
        incoming = (request.json or {}).get("contacts", [])
        if not isinstance(incoming, list):
            return jsonify({"error": "Bad payload."}), 400
        try:
            with open(CONTACTS_FILE, "w") as f:
                json.dump(incoming, f, indent=1)
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        return jsonify({"ok": True, "count": len(incoming)})
    try:
        with open(CONTACTS_FILE) as f:
            data = json.load(f)
            return jsonify({"contacts": data if isinstance(data, list) else []})
    except Exception:
        return jsonify({"contacts": []})


PUBLISHED_FILE = os.path.join(GENERATED_DIR, "published.json")


@app.route("/published", methods=["GET", "POST"])
def published():
    """The 'live site' snapshot: best deal per car at the moment the user last pushed.
    The hub diffs current bests against this to flag better / new / expired deals."""
    if request.method == "POST":
        data = request.json or {}
        try:
            with open(PUBLISHED_FILE, "w") as f:
                json.dump(data, f, indent=1)
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        return jsonify({"ok": True})
    try:
        with open(PUBLISHED_FILE) as f:
            d = json.load(f)
            return jsonify(d if isinstance(d, dict) else {"snapshot": {}, "at": ""})
    except Exception:
        return jsonify({"snapshot": {}, "at": ""})


INVOICES_FILE = os.path.join(GENERATED_DIR, "invoices.json")


@app.route("/invoices", methods=["GET", "POST"])
def invoices():
    """Saved customer invoices + the auto-increment counter. The client owns the
    list and number; the server just persists it (volume-backed)."""
    if request.method == "POST":
        data = request.json or {}
        rows = data.get("invoices")
        store = {"invoices": rows if isinstance(rows, list) else [],
                 "counter": int(data.get("counter") or 0)}
        try:
            with open(INVOICES_FILE, "w") as f:
                json.dump(store, f, indent=1)
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        return jsonify({"ok": True, "count": len(store["invoices"])})
    try:
        with open(INVOICES_FILE) as f:
            d = json.load(f)
            return jsonify({"invoices": d.get("invoices", []), "counter": int(d.get("counter") or 0)})
    except Exception:
        return jsonify({"invoices": [], "counter": 0})


PROGRAMS_FILE = os.path.join(GENERATED_DIR, "programs.json")


@app.route("/desk-programs", methods=["GET", "POST"])
def desk_programs():
    """The Desking program library — saved manufacturer programs (residual, money
    factor, term, fees …) keyed by vehicle. The client owns the list; the server
    just persists it (volume-backed), same as deals/invoices."""
    if request.method == "POST":
        data = request.json or {}
        rows = data.get("programs")
        store = {"programs": rows if isinstance(rows, list) else []}
        try:
            with open(PROGRAMS_FILE, "w") as f:
                json.dump(store, f, indent=1)
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        return jsonify({"ok": True, "count": len(store["programs"])})
    try:
        with open(PROGRAMS_FILE) as f:
            d = json.load(f)
            return jsonify({"programs": d.get("programs", [])})
    except Exception:
        return jsonify({"programs": []})


@app.route("/detect-plate", methods=["POST"])
def detect_plate():
    """Use a vision model to locate the rear license plate. Returns its 4 corners as
    fractions of width/height — we composite the real Nova plate there, so the model
    only DETECTS (never draws), avoiding any logo garbling."""
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    image_data = (request.json or {}).get("image_data", "")
    if not image_data:
        return jsonify({"error": "Missing image."}), 400
    if "," not in image_data:
        image_data = "data:image/png;base64," + image_data
    try:
        client = _oai()
        resp = client.chat.completions.create(
            model="gpt-4.1",
            messages=[{"role": "user", "content": [
                {"type": "text", "text":
                    "Find the rear LICENSE PLATE in this car photo. Respond with ONLY strict JSON: "
                    '{"found": true|false, "corners": [[x,y],[x,y],[x,y],[x,y]]} where corners are the '
                    "plate's TOP-LEFT, TOP-RIGHT, BOTTOM-RIGHT, BOTTOM-LEFT as decimal fractions (0-1) of "
                    "image width and height. Tightly bound the plate itself (not the frame). If there is no "
                    "visible plate, return {\"found\": false}."},
                {"type": "image_url", "image_url": {"url": image_data}},
            ]}],
            response_format={"type": "json_object"},
            max_tokens=300,
        )
        data = json.loads(resp.choices[0].message.content or "{}")
        return jsonify({"found": bool(data.get("found")), "corners": data.get("corners")})
    except Exception as e:
        info = classify_error(e)
        print(f"[detect-plate] error: {info['reason']} | {info['detail']}")
        return jsonify({"error": info["detail"] or info["reason"], "reason": info["reason"]}), info["status"]


@app.route("/generate-background", methods=["POST"])
def generate_background():
    if not API_KEY or API_KEY == "sk-your-key-here":
        return jsonify({"error": "OPENAI_API_KEY not configured on the server."}), 500
    style = (request.json or {}).get("style", "studio")
    try:
        client = _oai()
        result = client.images.generate(
            model="gpt-image-1",
            prompt=bg_prompt(style),
            n=1,
            size="1024x1536",
            quality="high",
            output_format="png",
        )
        image_bytes = base64.b64decode(result.data[0].b64_json)
        filename = f"bg_{style}_{os.urandom(4).hex()}.png"
        with open(os.path.join(GENERATED_DIR, filename), "wb") as f:
            f.write(image_bytes)
        return jsonify({"success": True, "filename": filename})
    except Exception as e:
        info = classify_error(e)
        print(f"[background] error: {info['reason']} | {info['detail']}")
        return jsonify({"error": info["detail"] or info["reason"], "reason": info["reason"]}), info["status"]


@app.route("/delete", methods=["POST"])
def delete_image():
    name = (request.json or {}).get("filename", "")
    # path-safety: only allow deleting a plain .png inside GENERATED_DIR
    name = os.path.basename(name)
    if not name.endswith(".png"):
        return jsonify({"error": "Invalid filename."}), 400
    path = os.path.join(GENERATED_DIR, name)
    if os.path.exists(path):
        os.remove(path)
        return jsonify({"success": True, "deleted": name})
    return jsonify({"error": "File not found."}), 404


@app.route("/upload", methods=["POST"])
def upload_image():
    """Save a (background-removed) car image so it joins the gallery and persists."""
    data = request.json or {}
    image_data = data.get("image_data", "")
    vehicle = data.get("vehicle", "").strip()
    if not image_data:
        return jsonify({"error": "Missing image data."}), 400
    try:
        if "," in image_data:                      # strip a data: URL prefix if present
            image_data = image_data.split(",", 1)[1]
        image_bytes = base64.b64decode(image_data)
        base = car_slug(vehicle, "", "") or "uploaded_car"
        filename = f"{base}.png"
        with open(os.path.join(GENERATED_DIR, filename), "wb") as f:
            f.write(image_bytes)
        return jsonify({"success": True, "filename": filename})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---- Verify live site: fetch the public deals page so we can reconcile it ----
def _host_is_public(host):
    """True only if every address the host resolves to is a public IP — blocks
    SSRF to localhost / private networks / link-local / cloud metadata."""
    try:
        infos = socket.getaddrinfo(host, None)
    except Exception:
        return False
    for info in infos:
        try:
            addr = ipaddress.ip_address(info[4][0])
        except Exception:
            return False
        if (addr.is_private or addr.is_loopback or addr.is_link_local
                or addr.is_reserved or addr.is_multicast or addr.is_unspecified):
            return False
    return bool(infos)


def _extract_inertia_deals(body):
    """The Nova site is an Inertia.js app: the full deal dataset is embedded as
    JSON in the root element's data-page="..." attribute. Pull it straight out as
    structured rows (year/make/model/trim/monthly/das/term) so we don't have to
    AI-parse scraped text. Returns None if the payload isn't present/parseable."""
    m = re.search(r'data-page="([^"]*)"', body)
    if not m:
        return None
    try:
        data = json.loads(_html_unescape(m.group(1)))
    except Exception:
        return None
    cars = None
    try:
        cars = data["props"]["car_data"]["cars"]
    except Exception:
        cars = None
    if not isinstance(cars, list):                       # fall back to a generic search
        found = []
        def walk(o):
            if found:
                return
            if (isinstance(o, list) and o and isinstance(o[0], dict)
                    and ({"deal", "deals"} & set(o[0].keys()))
                    and ({"title", "year", "car_trim"} & set(o[0].keys()))):
                found.append(o); return
            if isinstance(o, dict):
                for v in o.values():
                    walk(v)
            elif isinstance(o, list):
                for v in o:
                    walk(v)
        walk(data)
        cars = found[0] if found else None
    if not isinstance(cars, list):
        return None

    def _mo(d):
        try:
            return float(str(d.get("monthly_payment") or 1e9).replace(",", ""))
        except Exception:
            return 1e9

    rows = []
    for c in cars:
        if not isinstance(c, dict):
            continue
        deal = c.get("deal")
        deals = [d for d in (c.get("deals") or []) if isinstance(d, dict)]
        if not isinstance(deal, dict) and deals:
            deal = sorted(deals, key=_mo)[0]             # cheapest advertised
        if not isinstance(deal, dict):
            continue                                     # listed car with no live deal — skip
        make = model = ""
        try:
            cm = c["car_trim"]["car_model"]
            model = cm.get("name") or ""
            make = (cm.get("car_make") or {}).get("name") or ""
        except Exception:
            pass
        year, title = c.get("year") or "", str(c.get("title") or "")
        if not (make and model) and title:               # parse "2026 Toyota Corolla"
            parts = title.split()
            if parts and parts[0].isdigit():
                year = year or parts[0]; parts = parts[1:]
            if parts:
                make = make or parts[0]
            if len(parts) > 1:
                model = model or " ".join(parts[1:])
        rows.append({
            "year": str(year or ""), "make": make, "model": model,
            "trim": c.get("subtitle") or "",
            "monthly": str(deal.get("monthly_payment") or ""),
            "das": str(deal.get("down_payment") or ""),
            "term": str(deal.get("lease_term_months") or ""),
            "slug": str(c.get("slug") or ""),
        })
    return rows or None


def _html_to_text(h):
    """Crude HTML -> readable text: drop scripts/styles, turn block tags into
    line breaks, strip the rest, decode entities, collapse whitespace."""
    h = re.sub(r'(?is)<(script|style|noscript|head|svg)[^>]*>.*?</\1>', ' ', h)
    h = re.sub(r'(?i)<br\s*/?>', '\n', h)
    h = re.sub(r'(?i)</(p|div|li|tr|h[1-6]|section|article)>', '\n', h)
    h = re.sub(r'(?s)<[^>]+>', ' ', h)
    h = _html_unescape(h)
    h = re.sub(r'[ \t ]+', ' ', h)
    h = re.sub(r'\n[ \t]*\n+', '\n', h)
    return h.strip()


@app.route("/verify-fetch", methods=["POST"])
def verify_fetch():
    """Fetch the client-facing deals page (server-side, so no CORS) and return its
    readable text for the Deal Hub to reconcile against. SSRF-guarded."""
    url = (request.json or {}).get("url", "").strip()
    if not url:
        return jsonify({"error": "Enter the live site URL first."}), 400
    if not re.match(r'^https?://', url, re.I):
        url = "https://" + url
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in ("http", "https") or not parsed.hostname:
        return jsonify({"error": "That doesn't look like a valid web address."}), 400
    if not _host_is_public(parsed.hostname):
        return jsonify({"error": "That address can't be fetched (private or unreachable host)."}), 400
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (NovaVerify)"})
        with urllib.request.urlopen(req, timeout=20) as r:
            ctype = (r.headers.get_content_type() or "").lower()
            raw = r.read(3_000_000)            # cap at ~3 MB
    except urllib.error.HTTPError as e:
        return jsonify({"error": f"The site returned HTTP {e.code}."}), 502
    except Exception as e:
        return jsonify({"error": f"Couldn't reach the site: {e}"}), 502
    body = raw.decode("utf-8", "ignore")
    rows = _extract_inertia_deals(body)                  # the Nova site embeds its deals as JSON
    if rows:
        return jsonify({"found": True, "rows": rows, "count": len(rows), "source": "live-data", "url": url})
    text = _html_to_text(body) if ("html" in ctype or "<" in body[:300]) else body
    if not text.strip():
        return jsonify({"found": False, "message": "The page loaded but had no readable text (it may be image-only or JS-rendered — use Paste instead)."})
    return jsonify({"found": True, "text": text[:60000], "url": url})


if __name__ == "__main__":
    from waitress import serve
    port = int(os.environ.get("PORT", 5050))
    if not API_KEY or API_KEY == "sk-your-key-here":
        print("WARNING: No API key set. Set OPENAI_API_KEY (env) or add it to .env.")
    print(f"Image generator running at http://localhost:{port}")
    serve(app, host="0.0.0.0", port=port, channel_timeout=300)
